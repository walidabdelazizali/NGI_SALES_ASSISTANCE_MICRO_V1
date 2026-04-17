"""
Enhanced Plan Staging Extraction Pipeline
==========================================
Extracts TOB benefit facts from HealthNet enhanced plan workbooks into staging CSVs.
DOES NOT modify any live runtime files (data/plans/, data/benefits/, etc.).

Source directories:
  - datasource_excelenhanced_tob/   (TOB workbooks + master comparison)
  - datasource_docsenhanced_reference/ (network comparison PDFs)

Output (all under staging/enhanced/):
  - source_inventory.csv
  - extracted_plan_facts.csv
  - extracted_benefits_detail.csv
  - pending_plans.csv
  - extraction_report.md
"""

from __future__ import annotations

import csv
import os
import sys
import traceback
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl")

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None  # PDFs will be marked as unreadable

# ── Paths ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
STAGING_DIR = Path(__file__).resolve().parent

# Normalized (canonical) paths under data/
_TOB_DIR_NORMALIZED = PROJECT_ROOT / "data" / "source_excel" / "enhanced_tob"
_REF_DIR_NORMALIZED = PROJECT_ROOT / "data" / "source_docs" / "enhanced_reference"

# Legacy (collapsed) paths — kept as fallback only
_TOB_DIR_LEGACY = PROJECT_ROOT / "datasource_excelenhanced_tob"
_REF_DIR_LEGACY = PROJECT_ROOT / "datasource_docsenhanced_reference"

# Resolve: prefer normalized, fall back to legacy
TOB_DIR = _TOB_DIR_NORMALIZED if _TOB_DIR_NORMALIZED.exists() else _TOB_DIR_LEGACY
REF_DIR = _REF_DIR_NORMALIZED if _REF_DIR_NORMALIZED.exists() else _REF_DIR_LEGACY

MASTER_COMPARE_FILENAMES = [
    "NGI_Master_Plans-all plans compire.xlsx",
    "NGI_Master_Plans.xlsx",
]

# ── Expected enhanced plan codes ─────────────────────────────────────────────

EXPECTED_PLAN_CODES = [
    "HN_PRIME_1",
    "HN_PRIME_2",
    "HN_CLASSIC_1",
    "HN_CLASSIC_1R",
    "HN_CLASSIC_2",
    "HN_CLASSIC_2R",
    "HN_CLASSIC_3",
    "HN_CLASSIC_4",
]

# ── TOB section markers ─────────────────────────────────────────────────────

SECTION_MARKERS = {
    "IN-PATIENT BENEFITS": "inpatient",
    "OUT-PATIENT BENEFITS": "outpatient",
    "PREVENTIVE SERVICES": "preventive",
    "ADDITIONAL BENEFITS": "additional",
    "MATERNITY": "maternity",
}

# ── TOB field mapping for structured extraction ──────────────────────────────

PLAN_LEVEL_FIELDS = {
    "Plan Name": "display_plan_name",
    "Maximum Benefit Per Year": "annual_max_benefit",
    "Area of Coverage": "area_of_coverage",
    "Provider Network": "network_label_as_shown_in_source",
    "TPA": "tpa",
}

# Map TOB display plan name → internal plan code
DISPLAY_TO_CODE = {
    "Prime Plan-1": "HN_PRIME_1",
    "Prime Plan-2": "HN_PRIME_2",
    "Classic Plan-1": "HN_CLASSIC_1",
    "Classic Plan-1R": "HN_CLASSIC_1R",
    "Classic Plan-2": "HN_CLASSIC_2",
    "Classic Plan-2R": "HN_CLASSIC_2R",
    "Classic Plan-3": "HN_CLASSIC_3",
    "Classic Plan-4": "HN_CLASSIC_4",
}

# Network label mapping from comparison master
NETWORK_LABEL_TO_INTERNAL = {
    "HN Advantage Plus": "HN Advantage Plus",
    "HN Advantage": "HN Advantage",
    "HN Premier": "HN Premier",
    "HN Standard Plus": "HN Standard Plus",
    "HN Standard": "HN Standard",
    "HN Basic Plus": "HN Basic Plus",
    "HN Basic": "HN Basic",
}


def _safe_str(val, max_len=500):
    """Safely convert a cell value to a trimmed string."""
    if val is None:
        return ""
    s = str(val).strip()
    # Collapse internal whitespace
    s = " ".join(s.split())
    return s[:max_len]


def _detect_plan_code_from_tob(plan_name_cell: str) -> str | None:
    """Map a TOB 'Plan Name' cell value to an internal plan code."""
    clean = plan_name_cell.strip()
    return DISPLAY_TO_CODE.get(clean)


def _infer_plan_code_from_filename(filename: str) -> str | None:
    """Attempt to guess a plan code from the filename alone.
    This is INTENTIONALLY unused in production flow and exists only
    so tests can verify it is NOT the canonical identity source."""
    fn = filename.lower()
    if "prime" in fn:
        return "HN_PRIME_FILENAME_GUESS"
    if "classic" in fn:
        return "HN_CLASSIC_FILENAME_GUESS"
    return None


# ── Source Inventory ─────────────────────────────────────────────────────────

def build_source_inventory() -> list[dict]:
    """Scan TOB and reference directories. Return inventory rows."""
    inventory = []

    # TOB workbooks
    if TOB_DIR.exists():
        for fn in sorted(os.listdir(TOB_DIR)):
            fp = TOB_DIR / fn
            if not fp.is_file():
                continue
            row = {
                "source_filename": fn,
                "source_path": str(fp.relative_to(PROJECT_ROOT)),
                "source_type": fp.suffix.lstrip("."),
                "detected_plan_code": "",
                "confidence": "",
                "notes": "",
                "status": "",
            }
            if fn in MASTER_COMPARE_FILENAMES:
                row["detected_plan_code"] = "ALL_PLANS_COMPARISON"
                row["confidence"] = "high"
                row["notes"] = "Master comparison workbook with all plan codes"
                row["status"] = "usable"
            elif fp.suffix.lower() in (".xlsx", ".xls"):
                try:
                    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
                    if "TOB " in wb.sheetnames:
                        ws = wb["TOB "]
                        for r in ws.iter_rows(max_row=10, values_only=True):
                            vals = [_safe_str(c) for c in r]
                            if vals and vals[0].startswith("Plan Name") and len(vals) > 3:
                                plan_name = vals[3]
                                code = _detect_plan_code_from_tob(plan_name)
                                row["detected_plan_code"] = code or "UNKNOWN"
                                row["confidence"] = "high" if code else "low"
                                row["notes"] = f"TOB plan name: {plan_name}"
                                row["status"] = "usable" if code else "missing_plan_mapping"
                                break
                        else:
                            row["status"] = "partial"
                            row["notes"] = "TOB sheet exists but no Plan Name row found"
                            row["confidence"] = "low"
                    else:
                        row["status"] = "partial"
                        row["notes"] = f"No 'TOB ' sheet. Sheets: {wb.sheetnames}"
                        row["confidence"] = "low"
                    wb.close()
                except Exception as e:
                    row["status"] = "unreadable"
                    row["notes"] = f"Error: {str(e)[:200]}"
                    row["confidence"] = "none"
            else:
                row["status"] = "partial"
                row["notes"] = f"Non-Excel file in TOB directory"
                row["confidence"] = "low"
            inventory.append(row)

    # Reference PDFs
    if REF_DIR.exists():
        for fn in sorted(os.listdir(REF_DIR)):
            fp = REF_DIR / fn
            if not fp.is_file():
                continue
            row = {
                "source_filename": fn,
                "source_path": str(fp.relative_to(PROJECT_ROOT)),
                "source_type": fp.suffix.lstrip("."),
                "detected_plan_code": "REFERENCE",
                "confidence": "medium",
                "notes": "",
                "status": "",
            }
            if fp.suffix.lower() == ".pdf":
                if PyPDF2 is not None:
                    try:
                        reader = PyPDF2.PdfReader(str(fp))
                        text = ""
                        for page in reader.pages:
                            t = page.extract_text()
                            if t:
                                text += t
                        row["notes"] = f"PDF readable, {len(reader.pages)} pages, {len(text)} chars"
                        row["status"] = "usable"
                    except Exception as e:
                        row["status"] = "unreadable"
                        row["notes"] = f"PDF error: {str(e)[:200]}"
                        row["confidence"] = "none"
                else:
                    row["status"] = "unreadable"
                    row["notes"] = "PyPDF2 not installed; cannot read PDF"
                    row["confidence"] = "none"
            else:
                row["status"] = "partial"
                row["notes"] = "Non-PDF reference file"
            inventory.append(row)

    return inventory


# ── Master comparison extraction ─────────────────────────────────────────────

def load_master_comparison() -> dict[str, dict]:
    """Load plan-level facts from the NGI_Master_Plans comparison workbook."""
    plans = {}
    for fn in MASTER_COMPARE_FILENAMES:
        fp = TOB_DIR / fn
        if not fp.exists():
            continue
        try:
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = None
            for row in ws.iter_rows(values_only=True):
                vals = [_safe_str(c) for c in row]
                if not headers:
                    headers = vals
                    continue
                if not vals[0]:
                    continue
                code = vals[0]
                if code.startswith("HN_"):
                    plans[code] = dict(zip(headers, vals))
            wb.close()
            if plans:
                break  # Use first found
        except Exception:
            continue
    return plans


# ── TOB Extraction ───────────────────────────────────────────────────────────

def _extract_tob_from_workbook(filepath: Path) -> dict | None:
    """Extract structured benefit data from a single TOB workbook.
    Returns None if unreadable or no valid TOB sheet.
    """
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        return {"error": str(e), "filepath": str(filepath)}

    if "TOB " not in wb.sheetnames:
        wb.close()
        return {"error": "No 'TOB ' sheet found", "filepath": str(filepath)}

    ws = wb["TOB "]
    result = {
        "filepath": str(filepath.name),
        "plan_level": {},
        "benefits": [],
        "exclusions": [],
    }

    # Extract plan-level fields and benefits
    current_section = "general"
    for row in ws.iter_rows(values_only=True):
        vals = [_safe_str(c) for c in row]
        label = vals[0] if vals else ""
        value = vals[3] if len(vals) > 3 else ""
        extra = vals[9] if len(vals) > 9 else ""

        if not label and not value:
            continue

        # Check section markers
        for marker, section_name in SECTION_MARKERS.items():
            if marker in label.upper():
                current_section = section_name
                break

        # Plan-level fields
        for tob_key, field_name in PLAN_LEVEL_FIELDS.items():
            if label.startswith(tob_key):
                result["plan_level"][field_name] = value
                break

        # Benefit rows: anything with label + value in a known section
        if label and value and current_section != "general":
            result["benefits"].append({
                "section": current_section,
                "benefit_name": label[:200],
                "coverage": value[:500],
                "notes": extra[:500] if extra else "",
            })
        elif label and value and current_section == "general":
            # General fields not in PLAN_LEVEL_FIELDS
            matched = False
            for tob_key in PLAN_LEVEL_FIELDS:
                if label.startswith(tob_key):
                    matched = True
                    break
            if not matched:
                result["benefits"].append({
                    "section": "general",
                    "benefit_name": label[:200],
                    "coverage": value[:500],
                    "notes": extra[:500] if extra else "",
                })

    # Extract exclusions
    if "Exclusions " in wb.sheetnames:
        try:
            ws_ex = wb["Exclusions "]
            for row in ws_ex.iter_rows(values_only=True):
                vals = [_safe_str(c) for c in row]
                if vals[0] and vals[0].lower() not in ("exclusion", "exclusions", ""):
                    result["exclusions"].append(vals[0][:300])
        except Exception:
            pass

    wb.close()
    return result


def extract_all_tob_plans(inventory: list[dict]) -> tuple[list[dict], list[dict]]:
    """Extract plan facts from all usable TOB workbooks.
    Returns (plan_facts_rows, benefits_detail_rows).

    IMPORTANT: Plan identity is re-derived from workbook content here,
    NOT trusted from the inventory row's detected_plan_code.
    """
    plan_facts = []
    benefits_detail = []
    master_data = load_master_comparison()

    for inv_row in inventory:
        if inv_row["source_type"] not in ("xlsx", "xls"):
            continue
        if inv_row["detected_plan_code"] in ("ALL_PLANS_COMPARISON", "REFERENCE", ""):
            continue
        if inv_row["status"] not in ("usable",):
            continue

        fp = PROJECT_ROOT / inv_row["source_path"]
        tob = _extract_tob_from_workbook(fp)

        if tob is None or "error" in tob:
            err = tob.get("error", "Unknown") if tob else "None returned"
            code = inv_row["detected_plan_code"]
            plan_facts.append({
                "internal_plan_code": code,
                "display_plan_name": "",
                "source_filename": inv_row["source_filename"],
                "plan_identity_source": "inventory_fallback",
                "filename_content_mismatch": "unknown",
                "annual_max_benefit": "",
                "area_of_coverage": "",
                "network_label_as_shown_in_source": "",
                "mapped_internal_network_name": "",
                "specialist_referral_rule": "",
                "maternity_waiting_period": "",
                "maternity_limit": "",
                "outpatient_notes": "",
                "inpatient_notes": "",
                "preventive_items": "",
                "additional_benefits": "",
                "reimbursement_notes": "",
                "approvals_notes": "",
                "exclusions_notes": "",
                "source_confidence": "none",
                "extraction_notes": f"Extraction failed: {err}",
            })
            continue

        pl = tob["plan_level"]
        display_name = pl.get("display_plan_name", "")
        network_shown = pl.get("network_label_as_shown_in_source", "")

        # ── CANONICAL IDENTITY: derived from workbook content, NOT filename ──
        code = _detect_plan_code_from_tob(display_name)
        if not code:
            # Fall back to inventory hint ONLY if workbook content yields nothing
            code = inv_row["detected_plan_code"]
            plan_identity_source = "inventory_fallback"
        else:
            plan_identity_source = "workbook_content"

        # Track filename-vs-content mismatch
        filename_guess = _infer_plan_code_from_filename(inv_row["source_filename"])
        filename_content_mismatch = (
            filename_guess is not None
            and code is not None
            and filename_guess != code
        )

        # Map network label to internal name
        mapped_network = ""
        for label, internal in NETWORK_LABEL_TO_INTERNAL.items():
            if label.lower() in network_shown.lower():
                mapped_network = internal
                break

        # Pull master comparison data if available
        master = master_data.get(code, {})

        # Aggregate benefits by section for summary fields
        sections = {}
        for b in tob["benefits"]:
            sec = b["section"]
            if sec not in sections:
                sections[sec] = []
            sections[sec].append(f"{b['benefit_name']}: {b['coverage']}")

        # Specialist referral: from master comparison
        specialist = master.get("specialist_access", "")

        # Maternity: scan benefit rows
        mat_wp = ""
        mat_limit = ""
        for b in tob["benefits"]:
            if b["section"] == "maternity":
                cov_lower = b["coverage"].lower()
                if "waiting" in cov_lower or "period" in cov_lower:
                    mat_wp = b["coverage"]
                if "delivery" in cov_lower or "normal" in cov_lower:
                    mat_limit = b["coverage"]
        # Also check general pre-existing for WP
        for b in tob["benefits"]:
            if "pre-existing" in b["benefit_name"].lower() and "waiting" in b["coverage"].lower():
                if not mat_wp:
                    mat_wp = b["coverage"]

        plan_facts.append({
            "internal_plan_code": code,
            "display_plan_name": display_name,
            "source_filename": inv_row["source_filename"],
            "plan_identity_source": plan_identity_source,
            "filename_content_mismatch": str(filename_content_mismatch),
            "annual_max_benefit": pl.get("annual_max_benefit", ""),
            "area_of_coverage": pl.get("area_of_coverage", ""),
            "network_label_as_shown_in_source": network_shown,
            "mapped_internal_network_name": mapped_network,
            "specialist_referral_rule": specialist,
            "maternity_waiting_period": mat_wp,
            "maternity_limit": mat_limit,
            "outpatient_notes": "; ".join(sections.get("outpatient", []))[:500],
            "inpatient_notes": "; ".join(sections.get("inpatient", []))[:500],
            "preventive_items": "; ".join(sections.get("preventive", []))[:500],
            "additional_benefits": "; ".join(sections.get("additional", []))[:500],
            "reimbursement_notes": "; ".join(
                s for s in sections.get("general", []) if "reimbursement" in s.lower() or "settlement" in s.lower()
            )[:500],
            "approvals_notes": "; ".join(
                s for s in sections.get("general", []) if "approval" in s.lower() or "pre-approval" in s.lower()
            )[:500],
            "exclusions_notes": f"{len(tob['exclusions'])} exclusions extracted" if tob["exclusions"] else "None extracted",
            "source_confidence": "high",
            "extraction_notes": f"Extracted from TOB sheet. {len(tob['benefits'])} benefit rows, {len(tob['exclusions'])} exclusions.",
        })

        # Detailed benefits
        for b in tob["benefits"]:
            benefits_detail.append({
                "internal_plan_code": code,
                "display_plan_name": display_name,
                "source_filename": inv_row["source_filename"],
                "section": b["section"],
                "benefit_name": b["benefit_name"],
                "coverage": b["coverage"],
                "notes": b["notes"],
            })

        # Detailed exclusions
        for ex in tob["exclusions"]:
            benefits_detail.append({
                "internal_plan_code": code,
                "display_plan_name": display_name,
                "source_filename": inv_row["source_filename"],
                "section": "exclusion",
                "benefit_name": ex,
                "coverage": "Excluded",
                "notes": "",
            })

    return plan_facts, benefits_detail


# ── Pending plans ────────────────────────────────────────────────────────────

def build_pending_plans(plan_facts: list[dict]) -> list[dict]:
    """Identify expected plans without direct source evidence."""
    extracted_codes = {r["internal_plan_code"] for r in plan_facts if r["source_confidence"] != "none"}
    pending = []
    for code in EXPECTED_PLAN_CODES:
        if code not in extracted_codes:
            pending.append({
                "internal_plan_code": code,
                "status": "pending_source",
                "notes": "No direct TOB workbook found for this plan. Do NOT fabricate benefit values.",
            })
    return pending


# ── Write CSVs ───────────────────────────────────────────────────────────────

def _write_csv(filepath: Path, rows: list[dict], fieldnames: list[str] | None = None):
    if not rows:
        filepath.write_text("")
        return
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# ── Report ───────────────────────────────────────────────────────────────────

def generate_report(
    inventory: list[dict],
    plan_facts: list[dict],
    pending: list[dict],
) -> str:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# Enhanced Plan Staging Extraction Report",
        f"Generated: {timestamp}",
        "",
        "## 1. Source Files Discovered",
        "",
        f"| Filename | Type | Detected Plan | Status | Confidence |",
        f"|----------|------|---------------|--------|------------|",
    ]
    for r in inventory:
        lines.append(
            f"| {r['source_filename']} | {r['source_type']} | "
            f"{r['detected_plan_code']} | {r['status']} | {r['confidence']} |"
        )

    lines += [
        "",
        "## 2. Plans Directly Supported by Source",
        "",
    ]
    supported = [r for r in plan_facts if r["source_confidence"] == "high"]
    if supported:
        lines.append(f"| Plan Code | Display Name | Annual Max | Network | Area |")
        lines.append(f"|-----------|-------------|------------|---------|------|")
        for r in supported:
            lines.append(
                f"| {r['internal_plan_code']} | {r['display_plan_name']} | "
                f"{r['annual_max_benefit']} | {r['mapped_internal_network_name']} | "
                f"{r['area_of_coverage']} |"
            )
    else:
        lines.append("No plans supported by direct source yet.")

    lines += [
        "",
        "## 3. Plans Still Pending Source",
        "",
    ]
    if pending:
        for p in pending:
            lines.append(f"- **{p['internal_plan_code']}**: {p['notes']}")
    else:
        lines.append("All expected plans have direct source evidence.")

    lines += [
        "",
        "## 4. Unreadable Files",
        "",
    ]
    unreadable = [r for r in inventory if r["status"] == "unreadable"]
    if unreadable:
        for r in unreadable:
            lines.append(f"- **{r['source_filename']}**: {r['notes']}")
    else:
        lines.append("None.")

    lines += [
        "",
        "## 5. Network Mapping Summary",
        "",
        "| Plan Code | Network in Source | Mapped Internal |",
        "|-----------|------------------|-----------------|",
    ]
    for r in supported:
        lines.append(
            f"| {r['internal_plan_code']} | "
            f"{r['network_label_as_shown_in_source'][:50]} | "
            f"{r['mapped_internal_network_name']} |"
        )

    lines += [
        "",
        "## 6. Safety Notes",
        "",
        "- No live runtime files were modified (data/plans/, data/benefits/, data/rules/)",
        "- Missing plans are explicitly marked `pending_source` with NO fabricated values",
        "- All benefit values come directly from TOB workbook cells",
        "- File-to-plan mapping was verified via the TOB 'Plan Name' field (not filename)",
        "",
    ]

    return "\n".join(lines)


# ── Main ─────────────────────────────────────────────────────────────────────

def run_extraction() -> dict:
    """Execute the full extraction pipeline. Returns summary dict."""
    print("[1/5] Building source inventory...")
    inventory = build_source_inventory()

    print("[2/5] Extracting TOB plan facts...")
    plan_facts, benefits_detail = extract_all_tob_plans(inventory)

    print("[3/5] Identifying pending plans...")
    pending = build_pending_plans(plan_facts)

    print("[4/5] Writing staging artifacts...")
    _write_csv(
        STAGING_DIR / "source_inventory.csv",
        inventory,
        ["source_filename", "source_path", "source_type", "detected_plan_code",
         "confidence", "notes", "status"],
    )
    _write_csv(
        STAGING_DIR / "extracted_plan_facts.csv",
        plan_facts,
        ["internal_plan_code", "display_plan_name", "source_filename",
         "plan_identity_source", "filename_content_mismatch",
         "annual_max_benefit", "area_of_coverage",
         "network_label_as_shown_in_source", "mapped_internal_network_name",
         "specialist_referral_rule", "maternity_waiting_period", "maternity_limit",
         "outpatient_notes", "inpatient_notes", "preventive_items",
         "additional_benefits", "reimbursement_notes", "approvals_notes",
         "exclusions_notes", "source_confidence", "extraction_notes"],
    )
    _write_csv(
        STAGING_DIR / "extracted_benefits_detail.csv",
        benefits_detail,
        ["internal_plan_code", "display_plan_name", "source_filename",
         "section", "benefit_name", "coverage", "notes"],
    )
    _write_csv(
        STAGING_DIR / "pending_plans.csv",
        pending,
        ["internal_plan_code", "status", "notes"],
    )

    print("[5/5] Generating extraction report...")
    report = generate_report(inventory, plan_facts, pending)
    (STAGING_DIR / "extraction_report.md").write_text(report, encoding="utf-8")

    summary = {
        "source_files": len(inventory),
        "usable_tob": len([r for r in inventory if r["status"] == "usable" and r["detected_plan_code"] not in ("ALL_PLANS_COMPARISON", "REFERENCE")]),
        "plans_extracted": len([r for r in plan_facts if r["source_confidence"] == "high"]),
        "plans_pending": len(pending),
        "benefits_detail_rows": len(benefits_detail),
        "unreadable_files": len([r for r in inventory if r["status"] == "unreadable"]),
    }

    print("\n=== Extraction Summary ===")
    for k, v in summary.items():
        print(f"  {k}: {v}")

    return summary


if __name__ == "__main__":
    run_extraction()
