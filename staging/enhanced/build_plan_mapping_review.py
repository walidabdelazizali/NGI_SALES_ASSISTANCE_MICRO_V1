"""
Generate the plan_mapping_review.csv artifact for human review.
Combines TOB workbook evidence, master comparison data, and reconciliation status.
"""
import csv
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required")

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
STAGING_DIR = Path(__file__).resolve().parent

TOB_DIR = PROJECT_ROOT / "data" / "source_excel" / "enhanced_tob"
if not TOB_DIR.exists():
    TOB_DIR = PROJECT_ROOT / "datasource_excelenhanced_tob"

DISPLAY_TO_CODE = {
    "Prime Plan-1": "HN_PRIME_1",
    "Prime Plan-2": "HN_PRIME_2",
    "Classic Plan-1": "HN_CLASSIC_1",
    "Classic Plan-1R": "HN_CLASSIC_1R",
    "Classic Plan-2": "HN_CLASSIC_2",
    "Classic Plan-2R": "HN_CLASSIC_2R",
    "Classic Plan-3": "HN_CLASSIC_3",
    "Classic Plan-4": "HN_CLASSIC_4",
}

CODE_TO_DISPLAY = {v: k for k, v in DISPLAY_TO_CODE.items()}

MASTER_FILES = [
    "NGI_Master_Plans-all plans compire.xlsx",
    "NGI_Master_Plans.xlsx",
]

ALL_EXPECTED = [
    "HN_PRIME_1", "HN_PRIME_2",
    "HN_CLASSIC_1", "HN_CLASSIC_1R",
    "HN_CLASSIC_2", "HN_CLASSIC_2R",
    "HN_CLASSIC_3", "HN_CLASSIC_4",
]


def safe_str(val, max_len=500):
    if val is None:
        return ""
    return " ".join(str(val).strip().split())[:max_len]


def load_master():
    plans = {}
    for fn in MASTER_FILES:
        fp = TOB_DIR / fn
        if not fp.exists():
            continue
        try:
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = None
            for row in ws.iter_rows(values_only=True):
                vals = [safe_str(c) for c in row]
                if not headers:
                    headers = vals
                    continue
                if not vals[0]:
                    continue
                plans[vals[0]] = dict(zip(headers, vals))
            wb.close()
            if plans:
                break
        except Exception:
            continue
    return plans


def scan_tob_workbooks():
    results = []
    if not TOB_DIR.exists():
        return results
    for fn in sorted(os.listdir(TOB_DIR)):
        fp = TOB_DIR / fn
        if not fp.is_file() or fp.suffix.lower() not in (".xlsx", ".xls"):
            continue
        if fn.startswith("NGI_Master"):
            continue
        try:
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            if "TOB " not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["TOB "]
            plan_name = ""
            network = ""
            annual_max = ""
            for row in ws.iter_rows(max_row=15, values_only=True):
                vals = [safe_str(c) for c in row]
                label = vals[0] if vals else ""
                value = vals[3] if len(vals) > 3 else ""
                if label.startswith("Plan Name") and value:
                    plan_name = value
                if label.startswith("Provider Network") and value:
                    network = value
                if label.startswith("Maximum Benefit Per Year") and value:
                    annual_max = value
            wb.close()
            code = DISPLAY_TO_CODE.get(plan_name, "UNMAPPED")
            results.append({
                "source_file": fn,
                "workbook_plan_name": plan_name,
                "code": code,
                "workbook_network": network,
                "annual_limit": annual_max,
            })
        except Exception:
            continue
    return results


def build_review():
    tob_data = scan_tob_workbooks()
    master_data = load_master()

    tob_by_code = {}
    for t in tob_data:
        tob_by_code[t["code"]] = t

    rows = []
    seen = set()

    for t in tob_data:
        code = t["code"]
        seen.add(code)
        master = master_data.get(code, {})
        comp_network = master.get("network_level", "")
        comp_annual = master.get("annual_limit", "")

        is_prime = code.startswith("HN_PRIME")
        filename_misleading = (t["source_file"] != "" and t["workbook_plan_name"] != "")

        # Status determination
        if code == "UNMAPPED":
            status = "conflict"
            confidence = "none"
            notes = "Workbook plan name has no mapping entry"
        elif is_prime:
            if comp_network and t["workbook_network"]:
                if comp_network.lower() in t["workbook_network"].lower() or \
                   any(p.lower() in comp_network.lower() for p in t["workbook_network"].split() if len(p) > 2):
                    status = "confirmed"
                    confidence = "high"
                    notes = "Network evidence aligns"
                else:
                    status = "pending_mapping"
                    confidence = "medium"
                    notes = f"NETWORK CONFLICT: TOB='{t['workbook_network']}' vs Comparison='{comp_network}'"
            elif not comp_network:
                status = "pending_mapping"
                confidence = "medium"
                notes = "No comparison network for cross-check"
            else:
                status = "confirmed"
                confidence = "high"
                notes = "Identity from workbook content"
        else:
            status = "confirmed"
            confidence = "high"
            notes = "Classic plan confirmed from workbook content"

        if filename_misleading:
            notes += f" | FILENAME MISLEADING: '{t['source_file']}' != '{t['workbook_plan_name']}'"

        rows.append({
            "source_file": t["source_file"],
            "workbook_plan_name": t["workbook_plan_name"],
            "proposed_internal_code": code,
            "proposed_display_name": CODE_TO_DISPLAY.get(code, ""),
            "workbook_network": t["workbook_network"],
            "comparison_network": comp_network,
            "annual_limit": t["annual_limit"],
            "status": status,
            "confidence": confidence,
            "review_notes": notes,
        })

    # Expected plans without TOB workbooks
    for code in ALL_EXPECTED:
        if code in seen:
            continue
        master = master_data.get(code, {})
        rows.append({
            "source_file": "",
            "workbook_plan_name": "",
            "proposed_internal_code": code,
            "proposed_display_name": CODE_TO_DISPLAY.get(code, ""),
            "workbook_network": "",
            "comparison_network": master.get("network_level", ""),
            "annual_limit": master.get("annual_limit", ""),
            "status": "pending_source",
            "confidence": "none",
            "review_notes": "No TOB workbook found — do not fabricate",
        })

    return rows


def main():
    rows = build_review()

    fieldnames = [
        "source_file", "workbook_plan_name", "proposed_internal_code",
        "proposed_display_name", "workbook_network", "comparison_network",
        "annual_limit", "status", "confidence", "review_notes",
    ]
    fp = STAGING_DIR / "plan_mapping_review.csv"
    with open(fp, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Written: {fp}")

    print("\n=== Plan Mapping Review ===")
    for r in rows:
        print(f"  {r['proposed_internal_code']:20s} | {r['status']:18s} | {r['confidence']:6s} | {r['source_file']}")

    confirmed = [r for r in rows if r["status"] == "confirmed"]
    pending = [r for r in rows if r["status"] in ("pending_mapping", "pending_source")]
    conflict = [r for r in rows if r["status"] == "conflict"]
    print(f"\nConfirmed: {len(confirmed)}, Pending: {len(pending)}, Conflict: {len(conflict)}")


if __name__ == "__main__":
    main()
