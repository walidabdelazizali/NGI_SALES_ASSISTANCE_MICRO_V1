"""
Prime Reconciliation Script
Gathers all Prime-related evidence from TOB workbooks and master comparison,
then produces a reconciliation CSV artifact.
"""
import csv
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required")

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
STAGING_DIR = Path(__file__).resolve().parent

# Normalized paths
TOB_DIR = PROJECT_ROOT / "data" / "source_excel" / "enhanced_tob"
if not TOB_DIR.exists():
    TOB_DIR = PROJECT_ROOT / "datasource_excelenhanced_tob"

MASTER_FILES = [
    "NGI_Master_Plans-all plans compire.xlsx",
    "NGI_Master_Plans.xlsx",
]

DISPLAY_TO_CODE = {
    "Prime Plan-1": "HN_PRIME_1",
    "Prime Plan-2": "HN_PRIME_2",
    "Classic Plan-1": "HN_CLASSIC_1",
    "Classic Plan-1R": "HN_CLASSIC_1R",
    "Classic Plan-2": "HN_CLASSIC_2",
    "Classic Plan-2R": "HN_CLASSIC_2R",
    "Classic Plan-3": "HN_CLASSIC_3",
    "Classic Plan-4": "HN_CLASSIC_4",
}

ALL_EXPECTED_PRIME_CODES = ["HN_PRIME_1", "HN_PRIME_2"]
ALL_EXPECTED_CLASSIC_CODES = [
    "HN_CLASSIC_1", "HN_CLASSIC_1R",
    "HN_CLASSIC_2", "HN_CLASSIC_2R",
    "HN_CLASSIC_3", "HN_CLASSIC_4",
]


def safe_str(val, max_len=500):
    if val is None:
        return ""
    s = str(val).strip()
    return " ".join(s.split())[:max_len]


def get_tob_evidence():
    """Extract plan identity + network from each TOB workbook."""
    results = []
    if not TOB_DIR.exists():
        return results
    for fn in sorted(os.listdir(TOB_DIR)):
        fp = TOB_DIR / fn
        if not fp.is_file() or fp.suffix.lower() not in (".xlsx", ".xls"):
            continue
        if fn.startswith("NGI_Master"):
            continue
        try:
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            if "TOB " not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["TOB "]
            plan_name = ""
            network = ""
            annual_max = ""
            area = ""
            for row in ws.iter_rows(max_row=15, values_only=True):
                vals = [safe_str(c) for c in row]
                label = vals[0] if vals else ""
                value = vals[3] if len(vals) > 3 else ""
                if label.startswith("Plan Name") and value:
                    plan_name = value
                if label.startswith("Provider Network") and value:
                    network = value
                if label.startswith("Maximum Benefit Per Year") and value:
                    annual_max = value
                if label.startswith("Area of Coverage") and value:
                    area = value
            wb.close()
            code = DISPLAY_TO_CODE.get(plan_name, "UNMAPPED")
            results.append({
                "source_file": fn,
                "workbook_plan_name": plan_name,
                "inferred_internal_code": code,
                "workbook_network": network,
                "workbook_annual_max": annual_max,
                "workbook_area": area,
            })
        except Exception as e:
            results.append({
                "source_file": fn,
                "workbook_plan_name": f"ERROR: {e}",
                "inferred_internal_code": "ERROR",
                "workbook_network": "",
                "workbook_annual_max": "",
                "workbook_area": "",
            })
    return results


def get_master_evidence():
    """Extract plan facts from the master comparison workbook."""
    plans = {}
    for fn in MASTER_FILES:
        fp = TOB_DIR / fn
        if not fp.exists():
            continue
        try:
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = None
            for row in ws.iter_rows(values_only=True):
                vals = [safe_str(c) for c in row]
                if not headers:
                    headers = vals
                    continue
                if not vals[0]:
                    continue
                code = vals[0]
                plans[code] = dict(zip(headers, vals))
            wb.close()
            if plans:
                break
        except Exception:
            continue
    return plans


def reconcile():
    """Cross-reference TOB evidence with master comparison for all plans."""
    tob_evidence = get_tob_evidence()
    master_data = get_master_evidence()

    reconciliation = []

    # 1. Plans that have TOB workbooks
    seen_codes = set()
    for tob in tob_evidence:
        code = tob["inferred_internal_code"]
        seen_codes.add(code)
        master = master_data.get(code, {})
        master_network = master.get("network_level", "")
        master_annual = master.get("annual_limit", "")

        # Determine status
        is_prime = code.startswith("HN_PRIME")
        wb_network = tob["workbook_network"]

        if code == "UNMAPPED" or code == "ERROR":
            status = "conflict"
            confidence = "none"
            notes = "Could not map workbook plan name to internal code"
        elif is_prime:
            # Prime plans: check if network evidence aligns
            network_match = False
            if master_network and wb_network:
                # Check if master network label appears in workbook network
                if master_network.lower() in wb_network.lower():
                    network_match = True
                # Also check reverse
                for part in wb_network.split():
                    if part.lower() in master_network.lower():
                        network_match = True
                        break

            if network_match:
                status = "confirmed"
                confidence = "high"
                notes = "TOB workbook + master comparison agree on network"
            elif master_network and wb_network:
                # Both have network info but they don't match
                status = "pending_mapping"
                confidence = "medium"
                notes = f"Network mismatch: workbook='{wb_network}' vs master='{master_network}'"
            elif not master_network:
                status = "pending_mapping"
                confidence = "medium"
                notes = "No master comparison network to cross-reference"
            else:
                status = "confirmed"
                confidence = "high"
                notes = "TOB workbook content identity confirmed"
        else:
            # Classic plans with TOB source
            if master_network:
                status = "confirmed"
                confidence = "high"
                notes = "TOB workbook content + master comparison available"
            else:
                status = "confirmed"
                confidence = "high"
                notes = "TOB workbook content identity confirmed"

        reconciliation.append({
            "internal_plan_code": code,
            "workbook_display_name": tob["workbook_plan_name"],
            "source_file": tob["source_file"],
            "workbook_network": wb_network,
            "comparison_network": master_network,
            "workbook_annual_limit": tob["workbook_annual_max"],
            "comparison_annual_limit": master_annual,
            "status": status,
            "confidence": confidence,
            "review_notes": notes,
        })

    # 2. Expected plans without TOB workbooks
    all_expected = ALL_EXPECTED_PRIME_CODES + ALL_EXPECTED_CLASSIC_CODES
    for code in all_expected:
        if code in seen_codes:
            continue
        master = master_data.get(code, {})
        reconciliation.append({
            "internal_plan_code": code,
            "workbook_display_name": "",
            "source_file": "",
            "workbook_network": "",
            "comparison_network": master.get("network_level", ""),
            "workbook_annual_limit": "",
            "comparison_annual_limit": master.get("annual_limit", ""),
            "status": "pending_source",
            "confidence": "none",
            "review_notes": "No TOB workbook found. Do NOT fabricate.",
        })

    return reconciliation, tob_evidence, master_data


def write_reconciliation(reconciliation):
    """Write the reconciliation CSV."""
    fieldnames = [
        "internal_plan_code", "workbook_display_name", "source_file",
        "workbook_network", "comparison_network",
        "workbook_annual_limit", "comparison_annual_limit",
        "status", "confidence", "review_notes",
    ]
    fp = STAGING_DIR / "prime_reconciliation.csv"
    with open(fp, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(reconciliation)
    print(f"Written: {fp}")


if __name__ == "__main__":
    recon, tob_ev, master = reconcile()

    print("=== Prime Reconciliation ===")
    prime_rows = [r for r in recon if r["internal_plan_code"].startswith("HN_PRIME")]
    for r in prime_rows:
        print(f"\n  {r['internal_plan_code']}:")
        print(f"    Display: {r['workbook_display_name']}")
        print(f"    Source:  {r['source_file']}")
        print(f"    WB Net:  {r['workbook_network']}")
        print(f"    Cmp Net: {r['comparison_network']}")
        print(f"    Status:  {r['status']} ({r['confidence']})")
        print(f"    Notes:   {r['review_notes']}")

    print("\n=== All Plans Summary ===")
    for r in recon:
        print(f"  {r['internal_plan_code']:20s} | {r['status']:18s} | {r['confidence']:6s} | {r['workbook_display_name']}")

    write_reconciliation(recon)
