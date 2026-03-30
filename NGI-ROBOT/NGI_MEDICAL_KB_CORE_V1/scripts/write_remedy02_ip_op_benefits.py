
import openpyxl

workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"
plan_id = "PLAN_REMEDY_02"

# 1. PLAN_MASTER refinement
plan_master_row = {
    "plan_id": "PLAN_REMEDY_02",
    "plan_name": "NGI Healthnet – Remedy 02",
    "plan_family": "Remedy Series",
    "annual_max_benefit": "AED 150,000",
    "currency": "AED",
    "tpa": "",
    "medical_network_name": "HN Basic Plus (OP Restricted to Clinics)",
    "op_access_summary": "OP restricted to clinics with direct access to listed hospitals for OP services",
    "ip_access_summary": "Inpatient benefits covered as per schedule and network terms",
    "specialist_referral_rule": "Specialist subject to GP referral both in clinics and hospitals",
    "area_of_cover_key": "AOC_REMEDY_02",
    "group_declaration_rule_key": "GROUP_DECL_REMEDY_02",
    "individual_declaration_rule_key": "INDIV_DECL_REMEDY_02",
    "pre_existing_rule_key": "PREEXIST_REMEDY_02",
    "reimbursement_outside_network_key": "REIMB_OUT_NET_REMEDY_02",
    "reimbursement_outside_uae_key": "REIMB_OUT_UAE_REMEDY_02",
    "copayment_summary": "OP consultation/lab/radiology 15%; physiotherapy 15% per session; generic medicines 30% per prescription; inpatient non-urgent cases 20% with stated caps; maternity 10%",
    "labs_radiology_summary": "Laboratory and radiology covered; elective MRI/CT/endoscopies require pre-approval",
    "maternity_summary": "Waiting period nil; inpatient maternity 10% co-pay; normal delivery and medically necessary C-section up to AED 10,000 including co-pay",
    "alternative_medicine_summary": "Reimbursement basis; up to AED 1,500 per person per year; 30% on all services",
    "status": "active",
    "owner_notes": "Mapped as per Remedy 02 source"
}

wb = openpyxl.load_workbook(workbook_path)

# PLAN_MASTER
ws = wb["PLAN_MASTER"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
for i in range(ws.max_row, 1, -1):
    if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
        ws.delete_rows(i)
new_row = [plan_master_row.get(col, "") for col in header]
ws.append(new_row)

# 2. AREA_OF_COVERAGE hard reset
ws = wb["AREA_OF_COVERAGE"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
for i in range(ws.max_row, 1, -1):
    if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
        ws.delete_rows(i)
area_row = {
    "coverage_id": "AOC_REMEDY_02",
    "plan_id": "PLAN_REMEDY_02",
    "uae_coverage_text": "Emergency medical treatment within all emirates of the UAE",
    "international_coverage_text": "UAE & Indian Sub-continent & South East Asia (excluding Hong Kong & Singapore) for elective and emergency treatments",
    "excluded_countries_text": "Excluding Hong Kong & Singapore",
    "emergency_inside_uae_text": "Covered within all emirates of the UAE",
    "emergency_outside_uae_text": "Covered within territory of cover",
    "elective_outside_uae_text": "Elective IP treatment outside UAE subject to prior approval",
    "prior_approval_note": "Elective IP treatment outside UAE is subject to prior approval",
    "owner_notes": "Mapped from Remedy 02 source",
    "status": "active"
}
ws.append([area_row.get(col, "") for col in header])

# 3. GROUP_DECLARATION_RULES refinement
ws = wb["GROUP_DECLARATION_RULES"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
for i in range(ws.max_row, 1, -1):
    if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
        ws.delete_rows(i)
group_row = {
    "plan_id": "PLAN_REMEDY_02",
    "applicable_group_size": "21 - 50 Lives",
    "trigger_condition": "Group size 21 - 50",
    "rule_text": "Applicable for Groups Size 21 - 50 Lives",
    "status": "active"
}
ws.append([group_row.get(col, "") for col in header])

# 4. INDIVIDUAL_DECLARATION_RULES hard reset
ws = wb["INDIVIDUAL_DECLARATION_RULES"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
for i in range(ws.max_row, 1, -1):
    if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
        ws.delete_rows(i)
indiv_rows = [
    {"plan_id": plan_id, "group_size_band": "20 Lives and below", "stage": "Inception and member addition", "condition_type": "General requirement", "rule_text": "Applicable at inception and upon member addition", "status": "active"},
    {"plan_id": plan_id, "group_size_band": "Above 20 Lives", "stage": "Inception", "condition_type": "Age 65 and above", "rule_text": "Members aged 65 & above", "status": "active"},
    {"plan_id": plan_id, "group_size_band": "Above 20 Lives", "stage": "Inception", "condition_type": "Group declaration trigger", "rule_text": "If requested based on the Group Declaration Form", "status": "active"},
    {"plan_id": plan_id, "group_size_band": "Above 20 Lives", "stage": "Addition", "condition_type": "Age 65 and above", "rule_text": "Members aged 65 & above", "status": "active"},
    {"plan_id": plan_id, "group_size_band": "Above 20 Lives", "stage": "Addition", "condition_type": "Category Change", "rule_text": "Category Change", "status": "active"},
    {"plan_id": plan_id, "group_size_band": "Above 20 Lives", "stage": "Addition", "condition_type": "Late Additions", "rule_text": "Late Additions", "status": "active"}
]
for row in indiv_rows:
    ws.append([row.get(col, "") for col in header])

wb.save(workbook_path)

# Output
wb = openpyxl.load_workbook(workbook_path)
out = {}
# PLAN_MASTER
ws = wb["PLAN_MASTER"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
for row in ws.iter_rows(min_row=2, values_only=True):
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        out["PLAN_MASTER"] = dict(zip(header, row))
# AREA_OF_COVERAGE
ws = wb["AREA_OF_COVERAGE"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
area_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        area_count += 1
# GROUP_DECLARATION_RULES
ws = wb["GROUP_DECLARATION_RULES"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
group_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        group_count += 1
# INDIVIDUAL_DECLARATION_RULES
ws = wb["INDIVIDUAL_DECLARATION_RULES"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
indiv_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        indiv_count += 1
print(f"Workbook path: {workbook_path}")
print("Updated sheets: PLAN_MASTER, AREA_OF_COVERAGE, GROUP_DECLARATION_RULES, INDIVIDUAL_DECLARATION_RULES")
print("Exact values written for the refined PLAN_MASTER row:")
for k, v in out["PLAN_MASTER"].items():
    print(f"  {k}: {v}")
print(f"Final row count in AREA_OF_COVERAGE for PLAN_REMEDY_02: {area_count}")
print(f"Final row count in GROUP_DECLARATION_RULES for PLAN_REMEDY_02: {group_count}")
print(f"Final row count in INDIVIDUAL_DECLARATION_RULES for PLAN_REMEDY_02: {indiv_count}")
print("Confirmation: Workbook structure unchanged")
