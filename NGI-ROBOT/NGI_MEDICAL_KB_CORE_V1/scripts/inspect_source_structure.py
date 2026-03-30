import openpyxl
import os

source_xlsx = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\data\Medical Insurance Program – NGI Remedy - 02 Plan ررر.xlsx"

wb = openpyxl.load_workbook(source_xlsx, data_only=True)
sheet_names = wb.sheetnames

print(f"Source workbook: {source_xlsx}")
print("Sheet names:")
for s in sheet_names:
    print(f"  {s}")

for s in sheet_names:
    ws = wb[s]
    rows = list(ws.iter_rows(values_only=True))
    # Count used rows and columns
    used_rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
    used_row_count = len(used_rows)
    used_col_count = max((len([cell for cell in row if cell not in (None, "")]) for row in used_rows), default=0)
    print(f"\nSheet: {s}")
    print(f"  Used rows: {used_row_count}")
    print(f"  Used columns: {used_col_count}")
    print("  First 10 non-empty rows:")
    for row in used_rows[:10]:
        print("    ", row)

# Heuristic: Try to identify sheet purposes
print("\nSheet purpose guesses:")
for s in sheet_names:
    s_lower = s.lower()
    if "benefit" in s_lower:
        print(f"  {s}: benefits")
    elif "network" in s_lower and "provider" in s_lower:
        print(f"  {s}: network providers")
    elif "maternity" in s_lower:
        print(f"  {s}: maternity")
    elif "travel" in s_lower or "emergency" in s_lower:
        print(f"  {s}: travel emergency benefits")
    elif "declaration" in s_lower:
        print(f"  {s}: declaration rules")
    elif "pre_exist" in s_lower or "preexist" in s_lower:
        print(f"  {s}: pre-existing condition rules")
    else:
        print(f"  {s}: (purpose unclear)")
