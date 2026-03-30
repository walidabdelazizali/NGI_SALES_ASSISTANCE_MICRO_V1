# This script verifies the NGI_MEDICAL_KB_CORE_V1 Excel workbook for existence, sheet names, and header rows.
import os
from openpyxl import load_workbook

workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"
expected_sheets = [
    "README",
    "PLAN_MASTER",
    "AREA_OF_COVERAGE",
    "NETWORK_MASTER",
    "NETWORK_SPECIAL_PROVIDERS",
    "GROUP_DECLARATION_RULES",
    "INDIVIDUAL_DECLARATION_RULES",
    "PRE_EXISTING_RULES",
    "INPATIENT_BENEFITS",
    "OUTPATIENT_BENEFITS",
    "PREVENTIVE_VACCINES",
    "ADDITIONAL_BENEFITS",
    "TRAVEL_EMERGENCY_ASSIST",
    "MATERNITY",
    "ALIASES"
]

# 1. Check file existence
exists = os.path.isfile(workbook_path)
if not exists:
    print("Workbook file does not exist.")
    exit(1)

# 2. Try to open workbook
try:
    wb = load_workbook(workbook_path)
except Exception as e:
    print(f"Failed to open workbook: {e}")
    exit(1)

# 3. Check all sheet names
actual_sheets = wb.sheetnames
missing = [s for s in expected_sheets if s not in actual_sheets]
if missing:
    print(f"Missing sheets: {missing}")
    exit(1)

# 4. Check header row in every sheet
for name in expected_sheets:
    ws = wb[name]
    header = ws[1]
    if not header or not header[0].value:
        print(f"Missing header in sheet: {name}")
        exit(1)

print("VERIFIED: Workbook exists, opens, all sheets present, all headers present.")
