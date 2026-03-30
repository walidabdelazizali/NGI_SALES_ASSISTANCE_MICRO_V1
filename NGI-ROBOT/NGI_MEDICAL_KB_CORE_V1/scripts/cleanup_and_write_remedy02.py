import openpyxl
import docx
import os

workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"
source_docx = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\data\HN-REMEDY 2 .docx"
source_xlsx = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\data\Medical Insurance Program – NGI Remedy - 02 Plan ررر.xlsx"

wb = openpyxl.load_workbook(workbook_path)
sheet_names = wb.sheetnames
cleaned_sheets = []

# 1. Remove duplicated header rows from all sheets if row 2 duplicates row 1
def remove_duplicated_headers(ws):
    if ws.max_row >= 2:
        row1 = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
        row2 = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[2]]
        if row1 == row2:
            ws.delete_rows(2)
            return True
    return False

# 2. Delete all placeholder rows for Remedy 02 that contain incomplete values or notes
PLACEHOLDER_PHRASES = [
    "Remedy 02: Not mapped in this script.",
    "Filled from Remedy 02 sources on 2026-03-29"
]
def remove_placeholder_rows(ws):
    header = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
    plan_id_col = None
    notes_col = None
    for idx, col in enumerate(header):
        if col == "plan_id":
            plan_id_col = idx
        if col == "owner_notes":
            notes_col = idx
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if plan_id_col is not None and row[plan_id_col] == "PLAN_REMEDY_02":
            # Check for placeholder phrases in notes
            if notes_col is not None and row[notes_col]:
                for phrase in PLACEHOLDER_PHRASES:
                    if phrase in str(row[notes_col]):
                        rows_to_delete.append(i)
                        break
            # Check for incomplete rows (all empty except plan_id/notes)
            if notes_col is not None and all((cell == '' or cell is None) for j, cell in enumerate(row) if j not in [plan_id_col, notes_col]):
                rows_to_delete.append(i)
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)
    return len(rows_to_delete) > 0

for sheet in sheet_names:
    ws = wb[sheet]
    cleaned = False
    if remove_duplicated_headers(ws):
        cleaned = True
    if remove_placeholder_rows(ws):
        cleaned = True
    if cleaned:
        cleaned_sheets.append(sheet)

# Save after cleanup
wb.save(workbook_path)


# 3. Structured write for Remedy 02 (actual mapping logic)
import re
src_wb = openpyxl.load_workbook(source_xlsx, data_only=True)
doc = docx.Document(source_docx)
updated_sheets = []
rows_written = {}
plan_id = "PLAN_REMEDY_02"
network_id = "NET_HN_BASIC_PLUS_REMEDY_02"

# Helper: Find column index by name
def col_idx(header, name):
    for i, col in enumerate(header):
        if col and str(col).strip().lower() == name.lower():
            return i
    return None

# 3.1 PLAN_MASTER: Map plan-level info from xlsx (first sheet, first row)
if "PLAN_MASTER" in sheet_names and src_wb.sheetnames:
    ws = wb["PLAN_MASTER"]
    src_ws = src_wb[src_wb.sheetnames[0]]
    header = [cell.value for cell in ws[1]]
    src_row = [cell.value for cell in next(src_ws.iter_rows(min_row=2, max_row=2))]
    new_row = []
    for col in header:
        if col == "plan_id":
            new_row.append(plan_id)
        elif col == "network_id":
            new_row.append(network_id)
        elif col == "status":
            new_row.append("active")
        elif col == "owner_notes":
            new_row.append("")
        else:
            # Try to map from source row by header name
            idx = col_idx([cell.value for cell in src_ws[1]], col)
            new_row.append(src_row[idx] if idx is not None else "")
    ws.append(new_row)
    updated_sheets.append("PLAN_MASTER")
    rows_written["PLAN_MASTER"] = 1

# 3.2 NETWORK_SPECIAL_PROVIDERS: One provider per row from xlsx sheet if exists
if "NETWORK_SPECIAL_PROVIDERS" in sheet_names and "NETWORK_SPECIAL_PROVIDERS" in src_wb.sheetnames:
    ws = wb["NETWORK_SPECIAL_PROVIDERS"]
    src_ws = src_wb["NETWORK_SPECIAL_PROVIDERS"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("NETWORK_SPECIAL_PROVIDERS")
        rows_written["NETWORK_SPECIAL_PROVIDERS"] = count

# 3.3 INPATIENT_BENEFITS: One benefit per row from xlsx sheet if exists
if "INPATIENT_BENEFITS" in sheet_names and "INPATIENT_BENEFITS" in src_wb.sheetnames:
    ws = wb["INPATIENT_BENEFITS"]
    src_ws = src_wb["INPATIENT_BENEFITS"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("INPATIENT_BENEFITS")
        rows_written["INPATIENT_BENEFITS"] = count

# 3.4 OUTPATIENT_BENEFITS: One benefit per row from xlsx sheet if exists
if "OUTPATIENT_BENEFITS" in sheet_names and "OUTPATIENT_BENEFITS" in src_wb.sheetnames:
    ws = wb["OUTPATIENT_BENEFITS"]
    src_ws = src_wb["OUTPATIENT_BENEFITS"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("OUTPATIENT_BENEFITS")
        rows_written["OUTPATIENT_BENEFITS"] = count

# 3.5 GROUP_DECLARATION_RULES: One rule per row from xlsx sheet if exists
if "GROUP_DECLARATION_RULES" in sheet_names and "GROUP_DECLARATION_RULES" in src_wb.sheetnames:
    ws = wb["GROUP_DECLARATION_RULES"]
    src_ws = src_wb["GROUP_DECLARATION_RULES"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("GROUP_DECLARATION_RULES")
        rows_written["GROUP_DECLARATION_RULES"] = count

# 3.6 INDIVIDUAL_DECLARATION_RULES: One rule per row from xlsx sheet if exists
if "INDIVIDUAL_DECLARATION_RULES" in sheet_names and "INDIVIDUAL_DECLARATION_RULES" in src_wb.sheetnames:
    ws = wb["INDIVIDUAL_DECLARATION_RULES"]
    src_ws = src_wb["INDIVIDUAL_DECLARATION_RULES"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("INDIVIDUAL_DECLARATION_RULES")
        rows_written["INDIVIDUAL_DECLARATION_RULES"] = count

# 3.7 PRE_EXISTING_RULES: One rule per row from xlsx sheet if exists
if "PRE_EXISTING_RULES" in sheet_names and "PRE_EXISTING_RULES" in src_wb.sheetnames:
    ws = wb["PRE_EXISTING_RULES"]
    src_ws = src_wb["PRE_EXISTING_RULES"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("PRE_EXISTING_RULES")
        rows_written["PRE_EXISTING_RULES"] = count


# 3.8 AREA_OF_COVERAGE
if "AREA_OF_COVERAGE" in sheet_names and "AREA_OF_COVERAGE" in src_wb.sheetnames:
    ws = wb["AREA_OF_COVERAGE"]
    src_ws = src_wb["AREA_OF_COVERAGE"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("AREA_OF_COVERAGE")
        rows_written["AREA_OF_COVERAGE"] = count


# --- Remedy 02: NETWORK_MASTER and NETWORK_SPECIAL_PROVIDERS from vertical key-value source ---
if "NETWORK_MASTER" in sheet_names or "NETWORK_SPECIAL_PROVIDERS" in sheet_names:
    # Read Sheet1 as key-value pairs
    src_ws = src_wb[src_wb.sheetnames[0]]
    kv = {}
    for row in src_ws.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).strip():
            kv[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""

    # Parse Provider Network row
    provider_network = kv.get("Provider Network", "")
    # Example value: 'HN Basic Plus (OP Restricted to Clinics) with direct access to below hospitals for OP Services. ...\n•\tAster Hospitals (Qusais, Mankhool, Jebel Ali)\n•\tBurjeel Specialty Hospital Sharjah\n...'
    # Extract network name (before first parenthesis or before 'with')
    network_name = ""
    op_access_rule = ""
    referral_rule = ""
    provider_lines = []
    if provider_network:
        # Network name
        m = re.match(r"([^(\n]+?)(?:\s*\(| with|$)", provider_network)
        if m:
            network_name = m.group(1).strip()
        # OP access rule
        m2 = re.search(r"\(([^)]+)\)", provider_network)
        if m2:
            op_access_rule = m2.group(1).strip()
        # Referral rule (look for 'Specialist Subject to GP Referral'...)
        m3 = re.search(r"Specialist Subject to GP Referral[^\n\.]*", provider_network)
        if m3:
            referral_rule = m3.group(0).strip()
        # Provider lines (after 'below hospitals' or bullet points)
        provider_lines = re.findall(r"•\s*([^\n]+)", provider_network)

    # Write NETWORK_MASTER (one row)
    if "NETWORK_MASTER" in sheet_names:
        ws = wb["NETWORK_MASTER"]
        header = [cell.value for cell in ws[1]]
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "network_name":
                new_row.append(network_name)
            elif col == "op_access_rule":
                new_row.append(op_access_rule)
            elif col == "referral_rule":
                new_row.append(referral_rule)
            elif col == "status":
                new_row.append("active")
            else:
                new_row.append("")
        ws.append(new_row)
        updated_sheets.append("NETWORK_MASTER")
        rows_written["NETWORK_MASTER"] = 1

    # Write NETWORK_SPECIAL_PROVIDERS (one row per provider)
    if "NETWORK_SPECIAL_PROVIDERS" in sheet_names and provider_lines:
        ws = wb["NETWORK_SPECIAL_PROVIDERS"]
        header = [cell.value for cell in ws[1]]
        count = 0
        for provider in provider_lines:
            new_row = []
            for col in header:
                if col == "plan_id":
                    new_row.append(plan_id)
                elif col == "network_id":
                    new_row.append(network_id)
                elif col == "provider_name":
                    new_row.append(provider.strip())
                elif col == "status":
                    new_row.append("active")
                else:
                    new_row.append("")
            ws.append(new_row)
            count += 1
        updated_sheets.append("NETWORK_SPECIAL_PROVIDERS")
        rows_written["NETWORK_SPECIAL_PROVIDERS"] = count
        # Print provider names for output
        print("Providers written:")
        for provider in provider_lines:
            print(f"  {provider.strip()}")

# 3.10 PREVENTIVE_VACCINES
if "PREVENTIVE_VACCINES" in sheet_names and "PREVENTIVE_VACCINES" in src_wb.sheetnames:
    ws = wb["PREVENTIVE_VACCINES"]
    src_ws = src_wb["PREVENTIVE_VACCINES"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("PREVENTIVE_VACCINES")
        rows_written["PREVENTIVE_VACCINES"] = count

# 3.11 ADDITIONAL_BENEFITS
if "ADDITIONAL_BENEFITS" in sheet_names and "ADDITIONAL_BENEFITS" in src_wb.sheetnames:
    ws = wb["ADDITIONAL_BENEFITS"]
    src_ws = src_wb["ADDITIONAL_BENEFITS"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("ADDITIONAL_BENEFITS")
        rows_written["ADDITIONAL_BENEFITS"] = count

# 3.12 TRAVEL_EMERGENCY_ASSIST
if "TRAVEL_EMERGENCY_ASSIST" in sheet_names and "TRAVEL_EMERGENCY_ASSIST" in src_wb.sheetnames:
    ws = wb["TRAVEL_EMERGENCY_ASSIST"]
    src_ws = src_wb["TRAVEL_EMERGENCY_ASSIST"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("TRAVEL_EMERGENCY_ASSIST")
        rows_written["TRAVEL_EMERGENCY_ASSIST"] = count

# 3.13 MATERNITY
if "MATERNITY" in sheet_names and "MATERNITY" in src_wb.sheetnames:
    ws = wb["MATERNITY"]
    src_ws = src_wb["MATERNITY"]
    header = [cell.value for cell in ws[1]]
    count = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col == "status":
                new_row.append("active")
            elif col == "owner_notes":
                new_row.append("")
            else:
                idx = col_idx([cell.value for cell in src_ws[1]], col)
                new_row.append(row[idx] if idx is not None else "")
        ws.append(new_row)
        count += 1
    if count:
        updated_sheets.append("MATERNITY")
        rows_written["MATERNITY"] = count

# Save after writing (if any rows written)
wb.save(workbook_path)

# 4. Verification: check for duplicated headers and placeholder rows
post_cleaned = []
for sheet in sheet_names:
    ws = wb[sheet]
    if remove_duplicated_headers(ws) or remove_placeholder_rows(ws):
        post_cleaned.append(sheet)
wb.save(workbook_path)

print(f"Workbook path: {workbook_path}")
print("Cleaned sheets:")
for s in cleaned_sheets:
    print(f"  {s}")
print("Updated sheets:")
for s in updated_sheets:
    print(f"  {s}")
print("Rows written per sheet:")
for s, n in rows_written.items():
    print(f"  {s}: {n} row(s)")
if cleaned_sheets or post_cleaned:
    print("All duplicated headers and placeholder rows were removed.")
else:
    print("No duplicated headers or placeholder rows found.")
print("Remedy 02 rows should now contain only actual mapped values (if mapping logic is implemented).")

# Update NETWORK_SPECIAL_PROVIDERS for PLAN_REMEDY_02: split Aster Hospitals entry
ws = wb["NETWORK_SPECIAL_PROVIDERS"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
provider_col = header.index("provider_name") if "provider_name" in header else None
network_id_col = header.index("network_id") if "network_id" in header else None

# Collect all rows for PLAN_REMEDY_02
rows = list(ws.iter_rows(min_row=2, values_only=True))
rows_to_keep = []
for row in rows:
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        # Check if this is the combined Aster Hospitals row
        if provider_col is not None and row[provider_col] and "Aster Hospitals" in row[provider_col] and "," in row[provider_col]:
            # Split into three
            for loc in ["Qusais", "Mankhool", "Jebel Ali"]:
                new_row = list(row)
                new_row[provider_col] = f"Aster Hospitals - {loc}"
                rows_to_keep.append(tuple(new_row))
        else:
            rows_to_keep.append(row)
    else:
        rows_to_keep.append(row)

# Remove all PLAN_REMEDY_02 rows
for i in range(ws.max_row, 1, -1):
    if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
        ws.delete_rows(i)

# Append the split and kept rows
for row in rows_to_keep:
    ws.append(row)

# Save workbook
wb.save(workbook_path)

# Output result
# Count and list provider names for PLAN_REMEDY_02
provider_names = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        provider_names.append(row[provider_col])
print(f"Workbook path: {workbook_path}")
print(f"Total rows now present in NETWORK_SPECIAL_PROVIDERS for PLAN_REMEDY_02: {len(provider_names)}")
print("Exact provider names now present:")
for name in provider_names:
    print(f"  {name}")
