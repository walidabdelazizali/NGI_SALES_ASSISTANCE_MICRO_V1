import openpyxl
from docx import Document

workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"
source_xlsx = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\data\Medical Insurance Program – NGI Remedy - 02 Plan ررر.xlsx"

plan_id = "PLAN_REMEDY_02"
network_id = "NET_HN_BASIC_PLUS_REMEDY_02"

# Read key-value pairs from the source xlsx (vertical 2-column)
import re
src_wb = openpyxl.load_workbook(source_xlsx, data_only=True)
src_ws = src_wb[src_wb.sheetnames[0]]
kv = {}
for row in src_ws.iter_rows(values_only=True):
    if row and row[0] and str(row[0]).strip():
        kv[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""

group_decl = kv.get("Group Declaration form", "")
indiv_decl = kv.get("Individual Health Declaration form", "")
preexist = None
for k in kv:
    if k.lower().startswith("pre-existing"):
        preexist = kv[k]
        break

wb = openpyxl.load_workbook(workbook_path)

sheets_and_content = [
    ("GROUP_DECLARATION_RULES", group_decl),
    ("INDIVIDUAL_DECLARATION_RULES", indiv_decl),
    ("PRE_EXISTING_RULES", preexist)
]
rows_written = {}

for sheet, content in sheets_and_content:
    ws = wb[sheet]
    header = [cell.value for cell in ws[1]]
    plan_id_col = header.index("plan_id") if "plan_id" in header else None
    # 1. Delete all rows for PLAN_REMEDY_02
    for i in range(ws.max_row, 1, -1):
        if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
            ws.delete_rows(i)
    # 2. Write one rule per row (split by newlines or list markers)
    rules = []
    if content:
        # Try to split by numbered or bullet list, else by newlines
        if re.search(r"\d+\.\t", content):
            rules = re.split(r"\d+\.\t", content)
        elif '\n' in content:
            rules = [r.strip() for r in content.split('\n') if r.strip()]
        else:
            rules = [content.strip()]
    # Remove empty and duplicate rules
    rules = [r for r in rules if r]
    seen = set()
    unique_rules = []
    for r in rules:
        if r not in seen:
            unique_rules.append(r)
            seen.add(r)
    # Write each rule as a row
    count = 0
    for rule in unique_rules:
        new_row = []
        for col in header:
            if col == "plan_id":
                new_row.append(plan_id)
            elif col == "network_id":
                new_row.append(network_id)
            elif col.lower().startswith("rule") or col.lower().startswith("description"):
                new_row.append(rule)
            elif col == "status":
                new_row.append("active")
            else:
                new_row.append("")
        ws.append(new_row)
        count += 1
    rows_written[sheet] = count

wb.save(workbook_path)

# Reopen and verify
wb = openpyxl.load_workbook(workbook_path)
for sheet, _ in sheets_and_content:
    ws = wb[sheet]
    header = [cell.value for cell in ws[1]]
    plan_id_col = header.index("plan_id") if "plan_id" in header else None
    rows_written[sheet] = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if plan_id_col is not None and row[plan_id_col] == plan_id)

print(f"Workbook path: {workbook_path}")
print(f"Rows written to GROUP_DECLARATION_RULES: {rows_written['GROUP_DECLARATION_RULES']}")
print(f"Rows written to INDIVIDUAL_DECLARATION_RULES: {rows_written['INDIVIDUAL_DECLARATION_RULES']}")
print(f"Rows written to PRE_EXISTING_RULES: {rows_written['PRE_EXISTING_RULES']}")
print("Confirmed: These rows now exist for PLAN_REMEDY_02.")
