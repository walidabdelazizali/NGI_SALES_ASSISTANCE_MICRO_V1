import openpyxl

workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"
plan_id = "PLAN_REMEDY_02"
network_id = "NET_HN_BASIC_PLUS_REMEDY_02"
providers = [
    "Aster Hospitals - Qusais",
    "Aster Hospitals - Mankhool",
    "Aster Hospitals - Jebel Ali",
    "Burjeel Specialty Hospital Sharjah",
    "International Modern Hospital, DXB",
    "Ajman Specialty Hospital – Ajman",
    "Al Saha wa Al Shifaa Hospital For One Day Surgery – Sharjah"
]

# 1. Load the workbook
wb = openpyxl.load_workbook(workbook_path)
ws = wb["NETWORK_SPECIAL_PROVIDERS"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
provider_col = header.index("provider_name") if "provider_name" in header else None

# 2. Remove every data row where plan_id = PLAN_REMEDY_02
for i in range(ws.max_row, 1, -1):
    if ws.cell(row=i, column=plan_id_col+1).value == plan_id:
        ws.delete_rows(i)

# 3. Save the workbook
wb.save(workbook_path)

# 4. Reopen the workbook fresh
wb = openpyxl.load_workbook(workbook_path)
ws = wb["NETWORK_SPECIAL_PROVIDERS"]
header = [cell.value for cell in ws[1]]

# 5. Append exactly 7 provider rows
for provider in providers:
    new_row = []
    for col in header:
        if col == "plan_id":
            new_row.append(plan_id)
        elif col == "network_id":
            new_row.append(network_id)
        elif col == "provider_name":
            new_row.append(provider)
        elif col == "status":
            new_row.append("active")
        else:
            new_row.append("")
    ws.append(new_row)

# 6. Save again
wb.save(workbook_path)

# 7. Reopen again and count only rows where plan_id = PLAN_REMEDY_02
wb = openpyxl.load_workbook(workbook_path)
ws = wb["NETWORK_SPECIAL_PROVIDERS"]
header = [cell.value for cell in ws[1]]
plan_id_col = header.index("plan_id") if "plan_id" in header else None
provider_col = header.index("provider_name") if "provider_name" in header else None
final_providers = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if plan_id_col is not None and row[plan_id_col] == plan_id:
        final_providers.append(row[provider_col])

print(f"Workbook path: {workbook_path}")
print(f"Final count of NETWORK_SPECIAL_PROVIDERS rows for PLAN_REMEDY_02: {len(final_providers)}")
print("Exact provider names found after final reopen:")
for name in final_providers:
    print(f"  {name}")
if len(final_providers) == 7:
    print("Final count = 7. Success.")
else:
    print("Final count != 7. Check for issues.")
