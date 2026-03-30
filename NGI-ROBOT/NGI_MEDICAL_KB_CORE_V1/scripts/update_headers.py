# This script updates the header row in each sheet of the NGI_MEDICAL_KB_CORE_V1 Excel workbook.
import os
from openpyxl import load_workbook

workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"

headers = {
    "README": ["section", "key", "value", "owner_notes", "status"],
    "PLAN_MASTER": ["plan_id", "plan_name", "plan_family", "annual_max_benefit", "currency", "tpa", "medical_network_name", "op_access_summary", "ip_access_summary", "specialist_referral_rule", "area_of_cover_key", "group_declaration_rule_key", "individual_declaration_rule_key", "pre_existing_rule_key", "reimbursement_outside_network_key", "reimbursement_outside_uae_key", "copayment_summary", "labs_radiology_summary", "maternity_summary", "alternative_medicine_summary", "status", "owner_notes"],
    "AREA_OF_COVERAGE": ["coverage_id", "plan_id", "uae_coverage_text", "international_coverage_text", "excluded_countries_text", "emergency_inside_uae_text", "emergency_outside_uae_text", "elective_outside_uae_text", "prior_approval_note", "owner_notes", "status"],
    "NETWORK_MASTER": ["network_id", "plan_id", "network_name", "tpa", "op_general_rule", "ip_general_rule", "specialist_referral_rule", "owner_notes", "status"],
    "NETWORK_SPECIAL_PROVIDERS": ["provider_id", "plan_id", "network_id", "provider_name", "provider_group", "facility_type", "city", "emirate", "access_type", "service_scope", "inclusion_status", "notes", "owner_notes", "status"],
    "GROUP_DECLARATION_RULES": ["group_decl_rule_id", "plan_id", "applicable_group_size", "trigger_condition", "rule_text", "notes", "owner_notes", "status"],
    "INDIVIDUAL_DECLARATION_RULES": ["ind_decl_rule_id", "plan_id", "group_size_band", "stage", "condition_type", "rule_text", "notes", "owner_notes", "status"],
    "PRE_EXISTING_RULES": ["pre_existing_rule_id", "plan_id", "initial_exclusion_period", "emergency_exception_text", "uae_first_scheme_rule", "coverage_after_enrolment_rule", "hdf_requirement_text", "undeclared_condition_rule", "notes", "owner_notes", "status"],
    "INPATIENT_BENEFITS": ["ip_benefit_id", "plan_id", "benefit_name", "covered", "coverage_text", "annual_limit", "copayment", "approval_required", "special_conditions", "notes", "owner_notes", "status"],
    "OUTPATIENT_BENEFITS": ["op_benefit_id", "plan_id", "benefit_name", "covered", "coverage_text", "annual_limit", "copayment", "approval_required", "special_conditions", "notes", "owner_notes", "status"],
    "PREVENTIVE_VACCINES": ["preventive_benefit_id", "plan_id", "benefit_name", "covered", "coverage_text", "annual_limit", "copayment", "age_limit", "special_conditions", "notes", "owner_notes", "status"],
    "ADDITIONAL_BENEFITS": ["additional_benefit_id", "plan_id", "benefit_name", "covered", "coverage_text", "annual_limit", "copayment", "special_conditions", "notes", "owner_notes", "status"],
    "TRAVEL_EMERGENCY_ASSIST": ["travel_benefit_id", "plan_id", "benefit_name", "covered", "coverage_text", "annual_limit", "copayment", "territory_scope", "special_conditions", "notes", "owner_notes", "status"],
    "MATERNITY": ["maternity_benefit_id", "plan_id", "covered", "waiting_period", "annual_limit", "copayment", "normal_delivery_limit", "c_section_limit", "special_conditions", "notes", "owner_notes", "status"],
    "ALIASES": ["alias_id", "alias_type", "raw_value", "normalized_value", "target_entity_type", "target_entity_id", "target_entity_name", "owner_notes", "status"]
}

wb = load_workbook(workbook_path)
for sheet, header in headers.items():
    ws = wb[sheet]
    ws.delete_rows(1, 1)
    for col, value in enumerate(header, 1):
        ws.cell(row=1, column=col, value=value)
wb.save(workbook_path)
print("Workbook headers updated and written as cell values.")
