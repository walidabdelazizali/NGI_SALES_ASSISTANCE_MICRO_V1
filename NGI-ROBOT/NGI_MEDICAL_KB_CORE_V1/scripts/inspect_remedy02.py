import openpyxl
import os

# Target workbook path
workbook_path = r"C:\NGI-ROBOT\NGI_MEDICAL_KB_CORE_V1\workbook\NGI_MEDICAL_KB_CORE_V1.xlsx"

# Open workbook in read-only mode
wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

remedy02_sheets = {}
plan_id_value = "PLAN_REMEDY_02"

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    header = None
    plan_id_col = None
    remedy02_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(cell).strip() if cell is not None else '' for cell in row]
            # Try to find the plan_id column
            for idx, col_name in enumerate(header):
                if col_name.lower() == "plan_id":
                    plan_id_col = idx
                    break
            if plan_id_col is None:
                break  # No plan_id column, skip this sheet
            continue
        # Check for PLAN_REMEDY_02 in plan_id column
        if plan_id_col is not None and row[plan_id_col] == plan_id_value:
            remedy02_count += 1
    if remedy02_count > 0:
        remedy02_sheets[sheet_name] = remedy02_count

# Check for structure changes (basic check: sheetnames)
expected_sheets = [
    "PLAN_MASTER",
    "AREA_OF_COVERAGE",
    "NETWORK_MASTER",
    "NETWORK_SPECIAL_PROVIDERS",
    "GROUP_DECLARATION_RULES",
    "INDIVIDUAL_DECLARATION_RULES",
    "PRE_EXISTING_RULES",
    "INPATIENT_BENEFITS",
    "OUTPATIENT_BENEFITS",
    "PREVENTIVE_VACCINES",
    "ADDITIONAL_BENEFITS",
    "TRAVEL_EMERGENCY_ASSIST",
    "MATERNITY",
    "ALIASES"
]
structure_unchanged = (wb.sheetnames == expected_sheets)

print(f"Workbook path: {workbook_path}")
if remedy02_sheets:
    print("Sheets containing PLAN_REMEDY_02:")
    for sheet, count in remedy02_sheets.items():
        print(f"  {sheet}: {count} row(s)")
    print("Remedy 02 data exists in the workbook.")
else:
    print("No Remedy 02 data found in any sheet.")
print(f"Workbook structure unchanged: {structure_unchanged}")
print("No modifications were made to the workbook.")
