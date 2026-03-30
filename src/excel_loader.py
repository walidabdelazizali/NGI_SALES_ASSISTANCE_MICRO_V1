import openpyxl
from pathlib import Path

def load_workbook_data(workbook_path, sheet_names=None):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    data = {}
    for sheet in wb.sheetnames:
        if sheet_names and sheet not in sheet_names:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        data[sheet] = [dict(zip(header, row)) for row in rows[1:]]
    return data
