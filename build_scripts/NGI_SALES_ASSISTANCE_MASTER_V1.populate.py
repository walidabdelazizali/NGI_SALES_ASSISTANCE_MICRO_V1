import openpyxl
from openpyxl import Workbook

# Workbook and sheet definitions
wb = Workbook()

# Sheet: plan_master
plan_master_headers = [
    "plan_id", "plan_name", "product_type", "business_line", "plan_family", "plan_tier", "annual_max_benefit", "annual_max_benefit_currency", "area_of_coverage", "provider_network", "geographic_applicability", "source_document", "status", "notes"
]
ws_plan_master = wb.active
ws_plan_master.title = "plan_master"
ws_plan_master.append(plan_master_headers)
ws_plan_master.append([
    "REMEDY_02", "REMEDY 02", "Health Insurance", "Group", "Remedy", "02", "150000", "AED", "UAE & Indian Sub-continent & South EastAsia (Excl. HK & Singapore) for Elective & Emergency Treatments. Elective IP treatment outside UAE is subject to prior approval.", "HN Basic Plus (OP Restricted to Clinics) with direct access to below hospitals for OP Services. Aster Hospitals (Qusais, Mankhool, Jebel Ali)Burjeel Specialty Hospital SharjahInternational Modern Hospital, DXBAjman Specialty Hospital – AjmanAl Saha wa Al Shifaa Hospital For One Day Surgery – SharjahSpecialist Subject to GP Referral both in Clinics and Hospitals", "UAE, Indian Sub-continent, SE Asia", "HN-REMEDY 2 .docx", "active", "Seeded from validated Remedy 02 source"
])

# Sheet: plan_items_master
plan_items_master_headers = [
    "item_id", "plan_id", "section_type", "subsection", "item_name", "item_category", "item_summary", "coverage_status", "is_optional", "has_options", "default_value", "default_limit_value", "default_limit_unit", "approval_required", "approval_timing", "referral_required", "applicability_rule", "provider_scope", "comparison_value", "comparison_type", "display_order", "source_document", "notes"
]
ws_plan_items = wb.create_sheet("plan_items_master")
ws_plan_items.append(plan_items_master_headers)
# Example item (real population would iterate all items)
ws_plan_items.append([
    "RMD02-001", "REMEDY_02", "inpatient", "", "Tests, diagnosis, treatments and surgeries for Non-urgent cases", "fixed", "20% co-pay of max AED 500/incident and AED 1,000/year applied at payment. Above max co-pay, 100% covered.", "covered", False, False, "20%", "500/incident, 1000/year", "AED", True, "prior", True, "in_areas_of_coverage_only", "network", "20%", "copay", 1, "HN-REMEDY 2 .docx", "Seeded from validated Remedy 02 source"
])
# ... (repeat for all REMEDY_02 items)

# Sheet: plan_item_options
plan_item_options_headers = [
    "option_id", "item_id", "plan_id", "option_label", "option_type", "is_included", "member_cost_share", "limit_value", "limit_unit", "option_summary", "is_default", "display_order", "notes"
]
ws_plan_item_options = wb.create_sheet("plan_item_options")
ws_plan_item_options.append(plan_item_options_headers)
# Example option (real population would iterate all options)
ws_plan_item_options.append([
    "RMD02-001-OPT1", "RMD02-001", "REMEDY_02", "Standard", "fixed", True, "20%", "500/incident, 1000/year", "AED", "Standard co-pay option", True, 1, "Seeded from validated Remedy 02 source"
])
# ... (repeat for all options)

# Sheet: plan_comparison_matrix
plan_comparison_matrix_headers = [
    "comparison_id", "category", "item_name", "comparison_type", "REMEDY_01", "REMEDY_02", "REMEDY_03", "REMEDY_04", "REMEDY_05", "REMEDY_06", "notes"
]
ws_plan_comparison = wb.create_sheet("plan_comparison_matrix")
ws_plan_comparison.append(plan_comparison_matrix_headers)
ws_plan_comparison.append([
    "CMP-001", "inpatient", "Tests, diagnosis, treatments and surgeries for Non-urgent cases", "copay", "", "20% (max AED 500/incident, 1000/year)", "", "", "", "", "Seeded for REMEDY_02; others to be filled manually"
])
# ... (repeat for all comparison items)

# Save workbook
wb.save("NGI_SALES_ASSISTANCE_MASTER_V1.xlsx")
