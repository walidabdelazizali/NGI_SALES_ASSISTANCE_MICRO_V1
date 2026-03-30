import openpyxl
import json
import sys

SRC = 'NGI_SALES_ASSISTANCE_MICRO_V1/workbook/NGI_SALES_ASSISTANCE_MASTER_V1.xlsx'
SHEET = 'PLAN_MASTER'
OUT = 'NGI_SALES_ASSISTANCE_MICRO_V1/build_scripts/plan_master_extracted.json'

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]
    header = [cell.value for cell in ws[1]]
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2, values_only=True)]
    data = {'header': header, 'rows': rows}
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Extracted {len(rows)} rows from {SHEET} in {SRC}")
    print(f"Header: {header}")
    print(f"First 2 rows: {rows[:2]}")
    print(f"Output written to {OUT}")

if __name__ == '__main__':
    main()
