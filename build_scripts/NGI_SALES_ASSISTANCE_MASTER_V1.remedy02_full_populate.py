import openpyxl
import json
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('NGI_SALES_ASSISTANCE_MASTER_V1.xlsx')

# Load Remedy 02 source
with open('RAG_PROTOTYPE_BUILD_V1/source_plan_docs/structured/remedy_2_6_plan_master.json', encoding='utf-8') as f:
    plan_data = json.load(f)

# Helper: Find REMEDY_02 plan
remedy02 = None
for plan in plan_data['plans']:
    if plan['plan_id'] == 'REMEDY_02':
        remedy02 = plan
        break
if not remedy02:
    raise Exception('REMEDY_02 not found in source')

# --- plan_master ---
ws_master = wb['plan_master']
# Remove all but header
for row in ws_master.iter_rows(min_row=2, max_row=ws_master.max_row):
    for cell in row:
        cell.value = None
# Populate full REMEDY_02 row
plan_row = [
    remedy02.get('plan_id', ''),
    remedy02.get('plan_name', ''),
    'Health Insurance',
    'Group',
    'Remedy',
    '02',
    '150000',
    'AED',
    '', '', '', '', 'active', 'HN-REMEDY 2 .docx'
]
# Fill area_of_coverage, provider_network, etc. from Table of Benefits
for benefit in remedy02['sections']['Table of Benefits']:
    if benefit['benefit_key'] == 'area_of_coverage':
        plan_row[8] = benefit['values'][0]
    if benefit['benefit_key'] == 'provider_network':
        plan_row[9] = benefit['values'][0]
ws_master.append(plan_row)

# --- plan_items_master ---
ws_items = wb['plan_items_master']
# Remove all but header
for row in ws_items.iter_rows(min_row=2, max_row=ws_items.max_row):
    for cell in row:
        cell.value = None
item_id_counter = 1
section_map = {
    'Table of Benefits': 'plan_core',
    'Basic healthcare services for in-patients at authorized hospitalsReferral procedure: In respect of Essential benefit plan members, no costs incurred for advice, consultations or treatments provided by specialists or consultants without the insured first consulting a General Practitioner (or equivalent as designated by DHA) who is licensed by DHA or another competent UAE authority will be payable by the insurer. The GP must make his referral together with reasons via the DHA e-Referrals system (or other such temporary manual systems) for the claim to be considered.': 'inpatient',
    'Basic healthcare services for out-patients at authorized hospitals/clinics/health centersReferral procedure: In respect of Essential benefit plan members, no costs incurred for advice, consultations or treatments provided by specialists or consultants without the insured first consulting a General Practitioner (or equivalent as designated by DHA) who is licensed by DHA or another competent UAE authority will be payable by the insurer. The GP must make his referral together with reasons via the DHA e-Referrals system (or other such temporary manual systems) for the claim to be considered.': 'outpatient',
    'PREVENTIVE SERVICES, VACCINES AND IMMUNIZATIONS': 'preventive',
    'ADDITIONAL BENEFITS': 'additional',
    'NGI ISA ASSIST BENEFITS – FOR EMERGENCIES WHILE TRAVELING OUTSIDE COUNTRY OF RESIDENCE': 'isa_assist',
    'MATERNITY': 'maternity',
    'Summary of Premiums': 'terms_conditions',
}
for section, items in remedy02['sections'].items():
    section_type = section_map.get(section, 'out_of_scope')
    for benefit in items:
        item_id = f'RMD02-{item_id_counter:03d}'
        item_id_counter += 1
        ws_items.append([
            item_id,
            'REMEDY_02',
            section_type,
            '',
            benefit.get('label', ''),
            'fixed',
            '; '.join(benefit.get('values', [])),
            'covered',
            False,
            False,
            '', '', '', '', '', '', '', '', '', '', item_id_counter, 'HN-REMEDY 2 .docx', ''
        ])

# --- plan_item_options ---
ws_options = wb['plan_item_options']
# Remove all but header
for row in ws_options.iter_rows(min_row=2, max_row=ws_options.max_row):
    for cell in row:
        cell.value = None
# (For this plan, most items are fixed; if variants are found, add them here)
# Example: Maternity in-patient has co-pay and limits, so add as option
option_id = 1
for section, items in remedy02['sections'].items():
    for benefit in items:
        label = benefit.get('label', '').lower()
        if 'co-pay' in '; '.join(benefit.get('values', [])).lower() or 'coinsurance' in '; '.join(benefit.get('values', [])).lower():
            ws_options.append([
                f'RMD02-OPT{option_id:03d}',
                '',
                'REMEDY_02',
                benefit.get('label', ''),
                'optional_with_variants',
                True,
                '',
                '',
                '',
                '; '.join(benefit.get('values', [])),
                True,
                option_id,
                'HN-REMEDY 2 .docx'
            ])
            option_id += 1

# --- plan_comparison_matrix ---
ws_cmp = wb['plan_comparison_matrix']
# Remove all but header
for row in ws_cmp.iter_rows(min_row=2, max_row=ws_cmp.max_row):
    for cell in row:
        cell.value = None
cmp_id = 1
comparison_items = [
    ('annual_max_benefit', 'plan_core', 'Maximum Benefit Per Year'),
    ('area_of_coverage', 'plan_core', 'Area of Coverage'),
    ('provider_network', 'plan_core', 'Provider Network'),
    ('examination_and_treatment_by_authorized_general_practitioners_specialist_and_consultants', 'outpatient', 'Doctor Consultation'),
    ('laboratory_tests_carried_out_in_authorized_facility', 'outpatient', 'Diagnostics & Labs'),
    ('prescribed_drugs_and_medicines_generic', 'outpatient', 'Prescribed Drugs & Medicines'),
    ('dental_treatment', 'additional', 'Dental'),
    ('optical_benefits', 'additional', 'Optical'),
    ('mental_health', 'additional', 'Mental Health'),
    ('physiotherapy', 'outpatient', 'Physiotherapy'),
    ('organ_transplantation', 'additional', 'Organ Transplant'),
    ('kidney_dialysis', 'additional', 'Kidney Dialysis'),
    ('in_patient_maternity_services', 'maternity', 'Maternity'),
    ('newborn_cover', 'maternity', 'Newborn Cover'),
    ('essential_vaccinations_and_inoculations', 'preventive', 'Preventive Vaccines'),
]
for key, cat, name in comparison_items:
    value = ''
    for section, items in remedy02['sections'].items():
        for benefit in items:
            if benefit.get('benefit_key') == key:
                value = '; '.join(benefit.get('values', []))
    ws_cmp.append([
        f'RMD02-CMP{cmp_id:03d}',
        cat,
        name,
        '',
        '',
        value,
        '', '', '', '', '',
        'HN-REMEDY 2 .docx'
    ])
    cmp_id += 1

wb.save('NGI_SALES_ASSISTANCE_MASTER_V1.xlsx')
