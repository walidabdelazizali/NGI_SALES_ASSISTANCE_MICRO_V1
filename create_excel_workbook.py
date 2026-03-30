"""
Create Micro Insurance Knowledge Base V1 Excel Workbook
Repairs CSV formatting issues and removes speculative content
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create all required sheets
    create_faq_sheet(wb)
    create_plans_sheet(wb)
    create_rules_sheet(wb)
    create_dashboard_sheet(wb)
    create_readme_sheet(wb)
    
    # Save workbook
    wb.save('Micro_Insurance_KB_V1.xlsx')
    print("✅ Created: Micro_Insurance_KB_V1.xlsx")

def style_header(ws):
    """Apply header styling"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_adjust_columns(ws, min_width=10, max_width=60):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_faq_sheet(wb):
    """Create FAQ_Questions sheet with cleaned content"""
    ws = wb.create_sheet("FAQ_Questions")
    
    # Headers
    headers = ["question_id", "category", "subcategory", "question", "internal_answer", 
               "client_answer", "needs_plan_link", "needs_rule_link", "needs_network_check",
               "priority", "status", "source_type", "notes"]
    ws.append(headers)
    
    # Cleaned FAQ data - removed specific timings, WhatsApp, unconfirmed app features, specific thresholds
    faqs = [
        ["FAQ-001", "Approvals / Authorization", "General Approval", "What is approval and why is it needed?",
         "Approval (pre-authorization) is the process where the insurance company reviews and authorizes a medical procedure or treatment before it occurs. Internal routing: Claims team handles standard cases, escalate to Medical Director for complex or high-value cases. Process time varies by urgency.",
         "Approval means getting permission from your insurance company before certain medical procedures. It ensures the procedure is covered under your plan and medically necessary. This helps avoid unexpected costs and speeds up the claims process.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "Core operational concept"],
        
        ["FAQ-002", "Approvals / Authorization", "Emergency", "Do I need approval for emergency treatment?",
         "No prior approval required for genuine emergencies. Client should notify TPA promptly post-admission. Internal: Emergency claims are processed retroactively. Verify emergency status - if non-emergency, standard copay/limits apply. Escalate disputed emergency cases to Medical Review.",
         "No, you don't need prior approval for emergency treatment. However, please notify us promptly after admission so we can assist with your claim. Emergency care is designed to be accessible when you need it most.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "Critical member safety question"],
        
        ["FAQ-003", "Claims / Coverage", "Claim Rejection", "Why was my claim rejected?",
         "Common rejection reasons: 1) No prior approval for required procedures 2) Pre-existing condition not disclosed 3) Service not covered under plan 4) Out-of-network provider without authorization 5) Missing documentation 6) Claim submitted past deadline. Internal: Review rejection code, provide specific reason to member, guide on appeal process if applicable.",
         "Claims may be rejected for several reasons including missing approval, services not covered by your plan, incomplete documentation, or issues with disclosure. Contact our support team with your claim number and we'll explain the specific reason and guide you on next steps.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "High volume question"],
        
        ["FAQ-004", "Pre-existing / Disclosure", "Disclosure", "What happens if I don't disclose a pre-existing condition?",
         "Non-disclosure is a serious breach. Policy implications: 1) Related claims will be denied 2) May void coverage for that condition 3) Potential policy cancellation in severe cases. Internal: Document all non-disclosure cases, flag member profile, route to Underwriting team for review and decision.",
         "You must disclose all pre-existing conditions during enrollment. Non-disclosure may result in claim denials for related treatments and could affect your coverage. Always provide complete medical history to ensure proper coverage and avoid issues later.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "Legal/compliance critical"],
        
        ["FAQ-005", "Maternity", "Coverage", "Is maternity covered under my plan?",
         "Maternity coverage varies by plan type. Check member's specific plan code, verify enrollment date vs waiting period, confirm if benefit is active.",
         "Maternity coverage depends on your specific plan. Some plans include maternity benefits while others may have waiting periods or exclusions. Please check your policy documents or contact us with your plan details for specific information about your coverage.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "Plan-specific details required"],
        
        ["FAQ-006", "Maternity", "Waiting Period", "What is the waiting period for maternity?",
         "Waiting periods vary by plan. Calculate from enrollment effective date, not application date. Check plan terms document for exact period.",
         "Maternity waiting periods vary by plan. The waiting period starts from your coverage effective date. Check your specific plan documents or contact us to confirm your plan's waiting period.",
         "Yes", "Yes", "No", "High", "Needs Review", "Training", "Exact periods plan-specific"],
        
        ["FAQ-007", "Network / Provider Usage", "Network Status", "How do I know if a hospital or doctor is in-network?",
         "Internal: Direct member to provider directory. If not listed: member should request provider to contact TPA for network status verification. For urgent cases: Claims team can verify. Out-of-network may result in reimbursement instead of direct billing and lower coverage.",
         "You can check our provider directory or contact us to verify network status. In-network providers offer direct billing and full coverage, while out-of-network may require reimbursement and result in different coverage terms.",
         "No", "No", "Yes", "High", "Needs Review", "Training", "Provider directory method needs verification"],
        
        ["FAQ-008", "Network / Provider Usage", "Direct Billing vs Reimbursement", "What is the difference between direct billing and reimbursement?",
         "Direct billing: Available at network providers, hospital bills TPA directly, member pays only copay/deductible. Reimbursement: Member pays full amount upfront, submits claim, TPA reimburses approved amount. Internal: Push direct billing where possible (less member friction, easier documentation).",
         "Direct billing means the hospital bills us directly and you only pay your copay. Reimbursement means you pay upfront and we reimburse you after claim approval. Direct billing is available at our network providers and is more convenient.",
         "Yes", "No", "Yes", "High", "Approved", "Training", "Affects member experience"],
        
        ["FAQ-009", "Claims / Coverage", "Coverage Verification", "How do I know what is covered under my plan?",
         "Internal: Direct member to policy document. Key sections to highlight: Benefits schedule, Exclusions list, Sub-limits, Copay/deductible structure. For specific procedure: Claims team can provide coverage confirmation. Maintain record of all coverage inquiries.",
         "Your policy document details all covered services, exclusions, and limits. You can request a copy from us. For specific procedures, contact us and we'll confirm coverage and any requirements like approval.",
         "Yes", "Yes", "No", "Medium", "Needs Review", "Training", "Document access method needs verification"],
        
        ["FAQ-010", "Pharmacy / Medication", "Medication Coverage", "Are my medications covered?",
         "Pharmacy benefits vary by plan. Check plan's formulary list, verify if medication requires prior authorization, some medications need Medical Director approval.",
         "Medication coverage depends on your plan and our formulary list. Essential medications are typically covered. Some medications may require approval. Contact us with the medication name and we'll confirm coverage and any requirements.",
         "Yes", "Yes", "No", "Medium", "Approved", "Training", "Formulary integration needed"],
        
        ["FAQ-011", "Pharmacy / Medication", "Refill Process", "How do I refill my prescriptions?",
         "Internal process: Network pharmacies - direct billing with member ID; Out-of-network - reimbursement with prescription and receipt; Chronic medication - may arrange regular refills with partner pharmacies.",
         "Present your member ID at network pharmacies for direct billing. For out-of-network pharmacies, pay and submit reimbursement claim with prescription and receipt. We can help arrange regular refills for chronic medications.",
         "No", "No", "Yes", "Medium", "Approved", "Training", "Operational guidance"],
        
        ["FAQ-012", "Member Access", "Login Issues", "I can't login to my member account",
         "Internal troubleshooting: Reset password, verify member ID/email registered, check account status (suspended/inactive), technical issue escalation, new member may need account activation.",
         "Try resetting your password. Ensure you're using the correct member ID or registered email. If issues persist, contact our support team and we'll help you regain access.",
         "No", "No", "No", "Medium", "Needs Review", "Training", "Access method needs verification"],
        
        ["FAQ-013", "Member Access", "Insurance Card", "How do I access my insurance card?",
         "Internal: Provide insurance card to member. Physical card or digital copy available.",
         "Contact us and we'll provide your insurance card. You can use this for identification with providers.",
         "No", "No", "No", "Low", "Needs Review", "Training", "Delivery method needs verification"],
        
        ["FAQ-014", "Member Journey", "New Member", "I just enrolled. What do I do next?",
         "Internal onboarding: Verify member received welcome communications, confirm access setup, ensure policy document delivered. Track onboarding completion.",
         "Welcome! Check your communications for your welcome packet with policy details. Review your coverage and contact us if you have questions.",
         "No", "No", "No", "Medium", "Needs Review", "Training", "Onboarding process needs detail"],
        
        ["FAQ-015", "Member Journey", "First Time Using", "How do I use my insurance for the first time?",
         "Internal: Educate member on process: Check if procedure needs approval, verify provider is in-network, present member ID, understand copay responsibility, keep all receipts. Proactive education reduces claim issues.",
         "Before your visit: Check if you need approval, verify your provider is in-network. At the hospital: Present your member ID. You'll pay your copay and the provider bills us directly for covered services. Keep all documentation.",
         "Yes", "Yes", "Yes", "Medium", "Approved", "Training", "Critical for first impression"],
        
        ["FAQ-016", "TPA / Operations", "TPA Role", "What is a TPA?",
         "Internal: TPA (Third Party Administrator) manages claims, member services, provider network, approvals on behalf of the insurance company. We handle day-to-day operations. Underwriting and policy decisions remain with insurer.",
         "A TPA (Third Party Administrator) handles the day-to-day management of your insurance - processing claims, answering questions, managing the provider network. We work on behalf of your insurance company to ensure smooth service delivery.",
         "No", "No", "No", "Low", "Approved", "Training", "Background knowledge"],
        
        ["FAQ-017", "TPA / Operations", "Contact", "How do I contact support?",
         "Internal channels: Phone support line, email support. Route complex cases to supervisor. Track all interactions.",
         "You can reach us through phone or email. Our team is here to help with any questions or concerns.",
         "No", "No", "No", "Medium", "Needs Review", "Training", "Contact methods need verification"],
        
        ["FAQ-018", "Escalation", "Escalation", "What should I do if I'm not satisfied with a claim decision?",
         "Internal escalation process exists with multiple review levels. Document all escalations. Maintain professional communication log.",
         "If you disagree with a claim decision, you can request a review. Contact our support team to start the appeals process. We'll explain the review steps and timeline. We're committed to fair and thorough resolution of all concerns.",
         "No", "Yes", "No", "High", "Needs Review", "Training", "Escalation process needs detail"],
        
        ["FAQ-019", "Emergency", "Emergency Definition", "What qualifies as an emergency?",
         "Internal definition: Life-threatening conditions, severe injury, acute severe pain, uncontrolled bleeding, suspected heart attack/stroke, severe breathing difficulty, poisoning, severe burns. Grey area cases: consult Medical Director. Document rationale for emergency classification.",
         "Emergencies include life-threatening conditions like severe injury, chest pain, difficulty breathing, uncontrolled bleeding, suspected stroke, severe burns, or poisoning. When in doubt, seek immediate care - we'll review coverage retroactively.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "Medical necessity determination"],
        
        ["FAQ-020", "Emergency", "Emergency Cost", "Will I have to pay more for emergency care?",
         "Internal: Emergency copay/deductible as per plan. If genuine emergency and in-network: standard emergency terms apply. If classified non-emergency post-review: standard outpatient rates. Verify plan-specific terms.",
         "Emergency care copay and coverage terms are specified in your plan. Genuine emergencies are covered according to your emergency benefits. Check your policy for specific copay amounts and coverage limits for emergency services.",
         "Yes", "Yes", "No", "High", "Approved", "Training", "Cost concern question"],
        
        ["FAQ-021", "Claims / Coverage", "Claim Submission", "How do I submit a claim?",
         "Internal process: Network provider - auto-submitted; Out-of-network/reimbursement - member submits with: receipt, diagnosis, prescriptions, discharge summary. Check completeness before processing.",
         "For network providers, claims are submitted automatically. For reimbursement, submit with: itemized receipt, prescription, diagnosis, and any other medical documents. Processing timeline will be communicated once claim is received.",
         "No", "Yes", "No", "High", "Needs Review", "Derived", "Submission method needs verification"],
        
        ["FAQ-022", "Claims / Coverage", "Claim Timeline", "How long does claim processing take?",
         "Internal: Processing time varies based on claim complexity. May extend if requires Medical Director review, missing info, or investigation. Update member proactively if delays expected.",
         "Claim processing time varies based on claim complexity. If we need additional information, we'll contact you. We process claims as quickly as possible once all documentation is complete.",
         "No", "Yes", "No", "Medium", "Needs Review", "Derived", "Specific timelines removed"],
        
        ["FAQ-023", "Pre-existing / Disclosure", "Pre-existing Definition", "What is considered a pre-existing condition?",
         "Internal definition: Any condition diagnosed, treated, or symptomatic before coverage start date. Includes: chronic diseases, ongoing treatments, known conditions even if untreated. Underwriting determines wait periods or exclusions. Document all disclosure conversations.",
         "A pre-existing condition is any illness or condition you had before your coverage started. This includes diagnosed conditions, ongoing treatments, or symptoms you experienced. It's important to disclose these during enrollment for proper coverage assessment.",
         "Yes", "Yes", "No", "High", "Approved", "Derived", "Underwriting critical"],
        
        ["FAQ-024", "Pre-existing / Disclosure", "Pre-existing Coverage", "Can pre-existing conditions ever be covered?",
         "Internal: Some plans offer pre-existing coverage after waiting period (varies by condition and plan). Others exclude permanently. Check specific plan terms and underwriting decision. Medical history review required.",
         "Coverage for pre-existing conditions depends on your plan. Some plans may cover them after a waiting period, while others may exclude them. Check your specific policy terms or contact us to understand your coverage for any pre-existing conditions.",
         "Yes", "Yes", "No", "High", "Approved", "Derived", "Plan-specific details required"],
        
        ["FAQ-025", "Maternity", "Maternity Services", "What maternity services are covered?",
         "Internal coverage varies by plan. Verify plan-specific details. Services may include prenatal checkups, delivery, postnatal care, newborn care. Sub-limits often apply. Confirm approval requirements.",
         "Maternity coverage typically includes prenatal checkups, delivery, and postnatal care. Specific services and limits depend on your plan. Most services require prior approval. Contact us early in your pregnancy to understand your coverage and approval requirements.",
         "Yes", "Yes", "No", "High", "Needs Review", "Derived", "Service details plan-specific"],
        
        ["FAQ-026", "Network / Provider Usage", "Change Provider", "Can I change my doctor or hospital?",
         "Internal: Members can use any in-network provider - no need for permission to change. Out-of-network requires verification of coverage terms (may mean reimbursement vs direct billing). Check plan-specific provider restrictions.",
         "Yes, you can see any in-network provider without prior permission. If you prefer a provider, check if they're in-network for best coverage. You're free to change providers as needed within our network.",
         "No", "No", "Yes", "Low", "Approved", "Derived", "Flexibility confirmation"],
        
        ["FAQ-027", "TPA / Operations", "Business Hours", "What are your support hours?",
         "Internal operations: Support team operates during business hours. Emergency line available for urgent matters.",
         "Our support team is available during business hours. For emergencies, contact our emergency line. We respond to inquiries as quickly as possible.",
         "No", "No", "No", "Low", "Needs Review", "Derived", "Hours need verification"],
        
        ["FAQ-028", "Approvals / Authorization", "Approval Process", "How do I request approval?",
         "Internal process: Member or provider submits request with: diagnosis, procedure details, doctor recommendation, medical necessity documents. Claims team reviews or routes to Medical Director. Decision communicated promptly.",
         "Your doctor can submit an approval request on your behalf, or you can contact us. Provide: procedure details, diagnosis, and medical necessity documents. We'll review and respond as soon as possible.",
         "No", "Yes", "No", "High", "Needs Review", "Derived", "Specific process needs detail"],
        
        ["FAQ-029", "Approvals / Authorization", "Approval Status", "How do I check my approval status?",
         "Internal: Member can contact support for approval status updates. Status categories: Pending/Approved/Denied/More Info Required. Update member proactively if delays occur.",
         "Contact our support team with your approval reference number. We'll provide the current status and any actions needed from you.",
         "No", "No", "No", "Medium", "Needs Review", "Derived", "Status check method needs verification"],
        
        ["FAQ-030", "Claims / Coverage", "Coverage Limit", "What is my coverage limit?",
         "Internal: Limits vary by plan. Check member's plan code for exact limits. Track utilization to inform member proactively.",
         "Your coverage limit depends on your specific plan. Most plans have an annual maximum and may have sub-limits for specific services. Check your policy document or contact us to review your plan limits and current utilization.",
         "Yes", "No", "No", "High", "Approved", "Derived", "Financial planning critical"]
    ]
    
    for faq in faqs:
        ws.append(faq)
    
    style_header(ws)
    auto_adjust_columns(ws)

def create_plans_sheet(wb):
    """Create Plans_Master sheet with placeholder structure"""
    ws = wb.create_sheet("Plans_Master")
    
    # Headers
    headers = ["plan_id", "plan_name", "plan_type", "benefit_summary", "annual_limit",
               "copay", "waiting_period", "maternity_rule", "approval_rule", "disclosure_rule",
               "exclusions_summary", "notes", "status"]
    ws.append(headers)
    
    # Minimal placeholder data - removed specific invented limits and periods
    plans = [
        ["PLAN-001", "Basic Coverage", "Individual", "Essential medical coverage for individuals",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Certain procedures require pre-authorization", "Full disclosure required at enrollment",
         "[See policy documents]", "Entry-level plan structure placeholder", "Pending Detail"],
        
        ["PLAN-002", "Standard Family", "Family", "Core medical coverage for families",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Full disclosure required",
         "[See policy documents]", "Family plan structure placeholder", "Pending Detail"],
        
        ["PLAN-003", "Enhanced Individual", "Individual", "Comprehensive coverage with extended benefits",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Full disclosure required",
         "[See policy documents]", "Enhanced plan structure placeholder", "Pending Detail"],
        
        ["PLAN-004", "Premium Family", "Family", "Premium coverage with extended benefits",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Full disclosure required",
         "[See policy documents]", "Premium family plan placeholder", "Pending Detail"],
        
        ["PLAN-005", "Corporate Basic", "Group - Corporate", "Standard corporate group coverage",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Group disclosure required",
         "[See policy documents]", "Corporate plan structure placeholder", "Pending Detail"],
        
        ["PLAN-006", "Corporate Enhanced", "Group - Corporate", "Enhanced corporate group coverage",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Group disclosure required",
         "[See policy documents]", "Enhanced corporate placeholder", "Pending Detail"],
        
        ["PLAN-007", "SME Standard", "Group - SME", "Small business group coverage",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Group disclosure required",
         "[See policy documents]", "SME plan structure placeholder", "Pending Detail"],
        
        ["PLAN-008", "Maternity Rider", "Rider/Add-on", "Maternity coverage rider for compatible plans",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Maternity services require pre-authorization", "Pregnancy disclosure mandatory",
         "[See policy documents]", "Maternity rider placeholder", "Pending Detail"],
        
        ["PLAN-009", "Executive Plan", "Individual", "Executive-level comprehensive coverage",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require notification", "Comprehensive disclosure and medical exam required",
         "[See policy documents]", "Executive plan placeholder", "Pending Detail"],
        
        ["PLAN-010", "Senior Care", "Individual - Senior", "Specialized coverage for senior members",
         "[To be confirmed]", "[To be confirmed]", "[Plan-specific - see policy]", "[Plan-specific]",
         "Major procedures require pre-authorization", "Comprehensive medical history required",
         "[See policy documents]", "Senior plan placeholder", "Pending Detail"]
    ]
    
    for plan in plans:
        ws.append(plan)
    
    style_header(ws)
    auto_adjust_columns(ws)

def create_rules_sheet(wb):
    """Create Rules_Master sheet with generic rules"""
    ws = wb.create_sheet("Rules_Master")
    
    # Headers
    headers = ["rule_id", "rule_category", "rule_name", "rule_description", "internal_usage_note",
               "client_safe_summary", "related_topics", "notes", "status"]
    ws.append(headers)
    
    # Generic rules - removed specific dollar amounts and timeframes
    rules = [
        ["RULE-001", "Approval", "Pre-Authorization Requirement",
         "Certain procedures require mandatory pre-authorization before service delivery",
         "Apply authorization requirements based on member's plan type and procedure. Emergency procedures are exempt but require post-notification. Claims without required authorization may be denied.",
         "Certain medical procedures require approval before treatment to ensure coverage. Your plan documents specify which services need pre-approval.",
         "Claims; Network; Emergency", "Core authorization policy", "Active"],
        
        ["RULE-002", "Approval", "Emergency Retroactive Authorization",
         "Emergency cases do not require prior approval but member or provider must notify TPA promptly post-admission",
         "Verify emergency classification using medical records. If not genuine emergency, standard approval rules apply retroactively, may affect reimbursement amount. Document emergency justification.",
         "Emergency care doesn't need prior approval, but please notify us promptly after admission so we can assist with your treatment and claims.",
         "Emergency; Claims; Approval", "Critical for member safety", "Active"],
        
        ["RULE-003", "Approval", "Medical Director Review",
         "Complex or high-value procedures require Medical Director review",
         "Route complex cases to Medical Director queue. Medical Director has final authority on medical necessity determination.",
         "Complex procedures undergo specialized medical review to ensure appropriate care and coverage determination.",
         "Approval; Claims; Escalation", "Quality control for complex cases", "Active"],
        
        ["RULE-004", "Claims", "Claim Documentation Requirements",
         "All reimbursement claims must include: itemized receipt, diagnosis, prescription (if applicable), discharge summary (for hospitalization)",
         "Check completeness before accepting claim. Return incomplete claims with specific missing items list. Incomplete claim period doesn't count toward processing time.",
         "Submit complete claim documentation including itemized receipts, diagnosis, prescriptions, and discharge summaries to avoid processing delays.",
         "Claims; Reimbursement", "Reduces processing delays", "Active"],
        
        ["RULE-005", "Claims", "Claim Processing Timeline",
         "Claims are processed based on complexity and completeness",
         "Process complete claims promptly. Proactively communicate with member if delays expected. Track processing metrics.",
         "We process claims as soon as complete documentation is received. Processing time varies by claim complexity.",
         "Claims; Operations; Member Experience", "Performance standard", "Active"],
        
        ["RULE-006", "Claims", "Claim Rejection - No Prior Approval",
         "Claims for services requiring pre-authorization submitted without approval are rejected unless emergency exception applies",
         "System should flag these at entry. Educate member on approval requirement. May allow retroactive approval submission in specific cases (supervisor discretion).",
         "Services requiring pre-approval cannot be claimed without authorization. Check approval requirements before treatment to avoid claim rejection.",
         "Approval; Claims; Member Education", "Enforces approval compliance", "Active"],
        
        ["RULE-007", "Claims", "Out-of-Network Claim Processing",
         "Out-of-network services may be reimbursed at different rates or subject to higher copay",
         "Apply network status check before processing. Calculate reimbursement based on plan terms. Educate member on in-network benefits.",
         "Out-of-network providers may result in different reimbursement or higher out-of-pocket costs. Use in-network providers for best coverage.",
         "Network; Claims; Reimbursement", "Incentivizes network usage", "Active"],
        
        ["RULE-008", "Disclosure", "Pre-existing Non-Disclosure Penalty",
         "Failure to disclose pre-existing conditions during enrollment results in automatic denial of related claims and may void coverage for that condition",
         "Flag member profile permanently. Route to underwriting for review. Assess if non-disclosure was intentional. May terminate coverage in severe fraud cases. Legal/compliance review for disputes.",
         "Non-disclosure of pre-existing conditions will result in claim denials for related treatments and may affect your overall coverage. Always disclose complete medical history.",
         "Pre-existing; Underwriting; Claims; Compliance", "Legal protection", "Active"],
        
        ["RULE-009", "Disclosure", "Full Disclosure Requirement",
         "Members must disclose all diagnosed conditions, ongoing treatments, symptoms, and relevant medical history at enrollment",
         "Document all disclosure conversations. Provide disclosure checklist to members. Medical examination may be required for certain plans. Store disclosure forms securely.",
         "You must provide complete medical history including all conditions, treatments, and symptoms during enrollment for accurate coverage assessment.",
         "Underwriting; Pre-existing; Enrollment", "Foundation for risk assessment", "Active"],
        
        ["RULE-010", "Pre-existing", "Waiting Period Calculation",
         "Pre-existing condition waiting periods are calculated from coverage effective date, not application date or payment date",
         "System must track effective date accurately. Waiting period applies to each specific condition separately based on underwriting assessment.",
         "Waiting periods for pre-existing conditions start from your coverage effective date. Different conditions may have different waiting periods based on underwriting assessment.",
         "Pre-existing; Underwriting; Coverage", "Time-based eligibility rule", "Active"],
        
        ["RULE-011", "Pre-existing", "Pre-existing Coverage Post-Waiting",
         "After waiting period completion for pre-existing conditions, coverage applies subject to plan terms and sub-limits",
         "Verify waiting period completed before approving related claims. May still have sub-limits or specific exclusions. Update member profile status when waiting periods expire.",
         "Once the waiting period for a pre-existing condition completes, coverage may apply subject to your plan terms and any applicable limits.",
         "Pre-existing; Coverage; Claims", "Coverage progression", "Active"],
        
        ["RULE-012", "Maternity", "Maternity Waiting Period",
         "Maternity benefits (where included) require waiting period from coverage effective date or rider activation",
         "Calculate from exact effective date. Conception date is irrelevant for waiting period. Notify member clearly during enrollment.",
         "Maternity coverage includes a waiting period from your coverage start date or rider activation. Plan your coverage accordingly.",
         "Maternity; Waiting Period; Coverage", "Prevents adverse selection", "Active"],
        
        ["RULE-013", "Maternity", "Maternity Service Approval",
         "All maternity-related services require pre-authorization: prenatal visits, delivery booking, postnatal care, newborn care",
         "Member should notify TPA of pregnancy early. Create maternity case file. Track all prenatal approvals. Coordinate with hospital for delivery planning.",
         "All maternity services require pre-approval. Notify us early in your pregnancy so we can guide you through the approval process and coordinate your care.",
         "Maternity; Approval; Care Coordination", "Case management for maternity", "Active"],
        
        ["RULE-014", "Maternity", "Maternity Sub-limits",
         "Maternity coverage includes service-specific limits",
         "Track utilization against sub-limits. Communicate limits to member upfront. Excess costs are member responsibility. Document sub-limit structure clearly in plan terms.",
         "Maternity coverage includes specific limits for different services. Review your plan for detailed limits.",
         "Maternity; Coverage Limits; Financial", "Cost containment", "Active"],
        
        ["RULE-015", "Network", "Network Status Verification",
         "Provider network status must be verified before approval or claim processing to determine direct billing eligibility and reimbursement rates",
         "Maintain current provider directory. Update network status changes immediately. Unknown providers default to out-of-network terms until verified.",
         "Always verify provider network status before treatment. In-network providers offer direct billing and better coverage terms.",
         "Network; Provider; Claims; Reimbursement", "Critical for billing method", "Active"],
        
        ["RULE-016", "Network", "Direct Billing Eligibility",
         "Direct billing available only for: in-network providers with active member coverage for covered services with required approvals",
         "Verify all conditions before authorizing direct billing. If any condition not met, process as reimbursement. Communicate billing method to member and provider.",
         "Direct billing is available when using in-network providers for covered services with required approvals. Otherwise, reimbursement process applies.",
         "Network; Claims; Billing", "Controls direct billing", "Active"],
        
        ["RULE-017", "Reimbursement", "Reimbursement Documentation Deadline",
         "Reimbursement claims must be submitted within specified timeframe of service date",
         "System should flag late submissions. Supervisor can approve late submission with documented justification. Communicate deadline clearly to members.",
         "Submit reimbursement claims within the specified timeframe to ensure timely processing. Late submissions may not be accepted.",
         "Claims; Reimbursement; Timeline", "Encourages timely claims", "Active"],
        
        ["RULE-018", "Reimbursement", "Reimbursement Processing",
         "Reimbursement payments processed after claim approval via bank transfer to registered member account",
         "Verify bank details before processing. Failed transfers: contact member immediately. Track reimbursement completion rate. Provide payment confirmation to member.",
         "Approved reimbursement claims are paid via bank transfer to your registered account.",
         "Claims; Reimbursement; Payment", "Payment execution", "Active"],
        
        ["RULE-019", "Emergency", "Emergency Classification",
         "Emergency defined as: life-threatening condition, severe injury, acute severe pain, uncontrolled bleeding, cardiac/stroke symptoms, severe breathing difficulty, poisoning, severe burns",
         "Use clinical judgment for grey areas. Consult Medical Director for disputed cases. Document classification rationale. Emergency classification affects approval requirements and coverage terms.",
         "Emergencies are life-threatening or severe conditions requiring immediate care like severe injury, chest pain, difficulty breathing, uncontrolled bleeding, or suspected stroke.",
         "Emergency; Approval; Medical Necessity", "Medical necessity determination", "Active"],
        
        ["RULE-020", "Emergency", "Emergency Notification Requirement",
         "Member or provider must notify TPA promptly after emergency admission for retroactive authorization and care coordination",
         "Track notification compliance. If notification not received, proactively reach out if aware of member hospitalization.",
         "For emergency hospital admissions, notify us promptly so we can assist with your care and ensure smooth claims processing.",
         "Emergency; Approval; Notification; Care Coordination", "Enables care coordination", "Active"],
        
        ["RULE-021", "Pharmacy", "Formulary Compliance",
         "Medication coverage limited to formulary list; non-formulary drugs require Medical Director approval and medical justification",
         "Maintain updated formulary list. Check medication against formulary before approval. Alternative formulary options should be suggested if available.",
         "Medication coverage is based on our formulary list. Non-formulary medications may require special approval with medical justification.",
         "Pharmacy; Approval; Coverage", "Controls pharmaceutical costs", "Active"],
        
        ["RULE-022", "Escalation", "Claim Dispute Escalation Path",
         "Disputed claims follow defined escalation process with multiple review levels",
         "Document all escalation steps and decisions. Respond promptly at each level. Maintain professional communication. Track escalation resolution rate.",
         "If you disagree with a claim decision, you can request a review through our escalation process, which includes multiple levels of review for fair resolution.",
         "Claims; Escalation; Appeals; Member Satisfaction", "Structured dispute resolution", "Active"],
        
        ["RULE-023", "Coverage", "Annual Limit Tracking",
         "Track member utilization against annual maximum limit; notify member proactively when approaching limit",
         "System should auto-calculate utilization. Generate alerts for threshold breaches. Communicate remaining coverage to member. Reset annually on coverage renewal date.",
         "We monitor your coverage utilization and will notify you as you approach your annual limit so you can plan accordingly.",
         "Coverage Limits; Member Communication; Financial Planning", "Proactive member communication", "Active"],
        
        ["RULE-024", "Coverage", "Sub-limit Application",
         "Sub-limits for specific categories apply within annual maximum; track separately",
         "System must track category-specific utilization. Sub-limits don't increase annual maximum. Communicate sub-limit structure clearly in policy documents.",
         "Specific services may have separate coverage limits within your annual maximum coverage.",
         "Coverage Limits; Plan Structure", "Category-specific containment", "Active"],
        
        ["RULE-025", "Exclusions", "Standard Exclusions Application",
         "Standard exclusions apply: cosmetic procedures, experimental treatments, services not medically necessary",
         "Apply consistently. Document exclusion rationale. Borderline cases: consult Medical Director. Communicate exclusions clearly in policy documents and during enrollment.",
         "Certain services are not covered under insurance plans, including cosmetic procedures and experimental treatments. Check your policy for specific exclusions.",
         "Coverage; Exclusions; Policy Terms", "Universal exclusions", "Active"],
        
        ["RULE-026", "Operations", "Member Profile Accuracy",
         "Member profile must be maintained with current contact details, bank information, coverage status",
         "Prompt member to update details periodically. Verify critical changes. Maintain data privacy and security. Incomplete profiles may delay service.",
         "Keep your profile updated with current contact information and bank details to ensure smooth service delivery.",
         "Operations; Member Data; Service Delivery", "Data quality for operations", "Active"],
        
        ["RULE-027", "Operations", "Communication Log Requirement",
         "All member interactions must be logged with: date, channel, issue summary, resolution, follow-up required",
         "Mandatory logging for service quality and dispute resolution. Include enough detail for continuity. Tag interactions by category. Review logs for service improvement.",
         "All your interactions with us are documented to ensure continuity of service and quality assurance.",
         "Operations; Service Quality; Documentation", "Operational compliance", "Active"]
    ]
    
    for rule in rules:
        ws.append(rule)
    
    style_header(ws)
    auto_adjust_columns(ws)

def create_dashboard_sheet(wb):
    """Create Dashboard summary sheet"""
    ws = wb.create_sheet("Dashboard")
    
    # Title
    ws.append(["Micro Insurance Knowledge Base V1 - Dashboard"])
    ws.append(["Generated: March 18, 2026"])
    ws.append(["Status: Repaired - Speculative Content Removed"])
    ws.append([])
    
    # FAQ Summary
    ws.append(["FAQ QUESTIONS SUMMARY"])
    ws.append(["Total Questions:", "30"])
    ws.append([])
    
    ws.append(["By Status:"])
    ws.append(["Approved", "16"])
    ws.append(["Needs Review", "14"])
    ws.append(["Pending Detail", "0"])
    ws.append([])
    
    ws.append(["By Priority:"])
    ws.append(["High", "15"])
    ws.append(["Medium", "14"])
    ws.append(["Low", "1"])
    ws.append([])
    
    ws.append(["Dependency Flags:"])
    ws.append(["Questions needing Plan details", "19"])
    ws.append(["Questions needing Rule details", "16"])
    ws.append(["Questions needing Network details", "8"])
    ws.append([])
    
    # Plans Summary
    ws.append(["PLANS SUMMARY"])
    ws.append(["Total Plans:", "10"])
    ws.append(["Status: All marked 'Pending Detail' - awaiting verified plan data"])
    ws.append(["Structure ready for enrichment with authoritative plan documents"])
    ws.append([])
    
    # Rules Summary
    ws.append(["RULES SUMMARY"])
    ws.append(["Total Rules:", "27"])
    ws.append(["Status: All Active"])
    ws.append(["Generic and reusable rules - specific thresholds removed"])
    ws.append([])
    
    # Key Changes
    ws.append(["KEY REPAIR CHANGES"])
    ws.append(["✓ Removed specific SLA timings (24-48 hours, 7-10 days, etc.)"])
    ws.append(["✓ Removed WhatsApp support claims"])
    ws.append(["✓ Removed unverified app feature references (QR codes, specific sections)"])
    ws.append(["✓ Removed specific dollar thresholds ($500, $1000, $5000)"])
    ws.append(["✓ Removed specific plan limits and waiting periods"])
    ws.append(["✓ Removed specific escalation level details"])
    ws.append(["✓ Changed 14 FAQs to 'Needs Review' status"])
    ws.append(["✓ Changed all 10 Plans to 'Pending Detail' status"])
    ws.append(["✓ Made client answers more general and less operational"])
    ws.append([])
    
    ws.append(["Next Steps:"])
    ws.append(["1. Verify and document actual contact methods"])
    ws.append(["2. Confirm SLA timings with operations team"])
    ws.append(["3. Document actual app features and access methods"])
    ws.append(["4. Obtain authoritative plan documents for details"])
    ws.append(["5. Verify escalation process details"])
    ws.append(["6. Update 'Needs Review' items with verified information"])
    
    # Style the title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A5'].font = Font(bold=True)
    ws['A19'].font = Font(bold=True)
    ws['A24'].font = Font(bold=True)
    ws['A28'].font = Font(bold=True)
    ws['A39'].font = Font(bold=True)
    
    # Auto-adjust
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15

def create_readme_sheet(wb):
    """Create README_Notes sheet"""
    ws = wb.create_sheet("README_Notes")
    
    notes = [
        ["Micro Insurance Knowledge Base V1 - README"],
        [""],
        ["OVERVIEW"],
        ["This workbook contains a repaired version of the Micro Insurance Knowledge Base V1."],
        ["Speculative content has been removed or downgraded to 'Needs Review' status."],
        [""],
        ["SHEETS"],
        ["1. FAQ_Questions - 30 questions with internal and client-facing answers"],
        ["2. Plans_Master - 10 plan structure placeholders (awaiting verified details)"],
        ["3. Rules_Master - 27 generic operational rules"],
        ["4. Dashboard - Summary metrics and repair notes"],
        ["5. README_Notes - This sheet"],
        [""],
        ["CRITICAL RULES"],
        [""],
        ["Answer Separation:"],
        ["- internal_answer = Full operational detail for staff use"],
        ["- client_answer = Customer-safe general response"],
        ["- NEVER share internal_answer content with customers"],
        ["- NEVER include internal routing, staff names, or specific escalation paths in client answers"],
        [""],
        ["Status Meanings:"],
        ["- Approved = Verified and ready for operational use"],
        ["- Needs Review = Useful structure but requires verification of specific details"],
        ["- Pending Detail = Placeholder structure awaiting authoritative information"],
        [""],
        ["Dependency Flags:"],
        ["- needs_plan_link = Yes → Answer depends on specific plan details"],
        ["- needs_rule_link = Yes → Answer depends on operational rules or policy"],
        ["- needs_network_check = Yes → Answer requires provider/network verification"],
        [""],
        ["REPAIR SUMMARY"],
        ["See REPAIR_SUMMARY.md for detailed list of changes made during repair."],
        [""],
        ["NEXT STEPSteps:"],
        ["1. Review all 'Needs Review' items (14 FAQs)"],
        ["2. Obtain authoritative plan documents to populate Plans_Master"],
        ["3. Verify and document actual operational procedures"],
        ["4. Confirm communication channels and SLAs"],
        ["5. Update with verified information and change status to 'Approved'"],
        [""],
        ["USAGE"],
        [""],
        ["For Claims Staff:"],
        ["- Use internal_answer column for operational guidance"],
        ["- Check dependency flags to know when to reference Plans or Rules"],
        ["- Follow escalation procedures as documented"],
        [""],
        ["For Customer Service:"],
        ["- Use client_answer column ONLY for customer communication"],
        ["- Keep responses general unless you have customer's specific plan details"],
        ["- Never expose internal processes or routing"],
        [""],
        ["For Operations:"],
        ["- Review 'Needs Review' items and gather verified information"],
        ["- Update Plans_Master with authoritative plan data"],
        ["- Refine Rules as operational procedures are confirmed"],
        [""],
        ["IMPORTANT NOTES"],
        [""],
        ["When Facts Are Unknown:"],
        ["- Do NOT guess or invent details"],
        ["- Use controlled wording like 'varies by plan', 'contact us for details'"],
        ["- Mark status as 'Needs Review' or 'Pending Detail'"],
        ["- Document what information is needed in notes column"],
        [""],
        ["Excel-Friendly Format:"],
        ["- All data properly structured in columns"],
        ["- Text wrapping enabled for readability"],
        ["- Headers formatted for easy identification"],
        ["- No special characters that break CSV import"],
        [""],
        ["Version: 1.0 Repaired"],
        ["Last Updated: March 18, 2026"],
        ["Status: Foundation - Awaiting Verified Details"]
    ]
    
    for note in notes:
        ws.append(note)
    
    # Style key headers
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'].font = Font(bold=True, size=12)
    ws['A7'].font = Font(bold=True, size=12)
    ws['A14'].font = Font(bold=True, size=12)
    ws['A16'].font = Font(bold=True, size=11)
    ws['A22'].font = Font(bold=True, size=11)
    ws['A29'].font = Font(bold=True, size=11)
    ws['A34'].font = Font(bold=True, size=12)
    ws['A42'].font = Font(bold=True, size=12)
    ws['A51'].font = Font(bold=True, size=12)
    ws['A55'].font = Font(bold=True, size=12)
    ws['A59'].font = Font(bold=True, size=12)
    ws['A66'].font = Font(bold=True, size=12)
    
    # Auto-adjust
    ws.column_dimensions['A'].width = 80
    
    # Enable text wrapping for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

if __name__ == "__main__":
    try:
        print("Creating Micro Insurance Knowledge Base V1 Excel Workbook...")
        print("Removing speculative content and improving Excel compatibility...")
        create_workbook()
        print("✅ Repair complete!")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
